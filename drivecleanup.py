#!/usr/bin/env python3
"""
Google Drive Cleanup Analyzer
Analyzes Google Drive content and suggests files/folders to delete.
"""

import os
import pickle
import sys
import io
import tempfile
import re
import glob
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from loguru import logger

# For single-key input
try:
    import termios
    import tty
    HAS_TERMIOS = True
except ImportError:
    HAS_TERMIOS = False

try:
    import msvcrt
    HAS_MSVCRT = True
except ImportError:
    HAS_MSVCRT = False

# Content extraction libraries
try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# AWS Bedrock for Claude
try:
    import boto3
    import json
    HAS_BEDROCK = True
except ImportError:
    HAS_BEDROCK = False

# OAuth Scopes
SCOPES_READONLY = ['https://www.googleapis.com/auth/drive.readonly']
SCOPES_WRITE = ['https://www.googleapis.com/auth/drive']  # For deletion

# Configure logger (stdout only initially, file logging added later)
logger.remove()
logger.add(
    sys.stdout,
    format="<green>{time:HH:mm:ss}</green> | <level>{level: <8}</level> | <level>{message}</level>",
    level="INFO",
    colorize=True
)

def setup_file_logging(folder_id, start_time):
    """Set up file logging for the session."""
    timestamp = start_time.strftime('%Y%m%d_%H%M%S')
    log_filename = f"{folder_id}_session_{timestamp}.log"

    # Add file handler
    logger.add(
        log_filename,
        format="{time:YYYY-MM-DD HH:mm:ss.SSS UTC} | {level: <8} | {message}",
        level="DEBUG",  # Log everything to file
        backtrace=True,
        diagnose=True
    )

    logger.info(f"Session log file: {log_filename}")
    return log_filename

# ============================================================================
# ANALYSIS CRITERIA
# ============================================================================

# File patterns that are typically safe to delete
TEMP_PATTERNS = [
    'tmp', 'temp', 'cache', '.cache', '.tmp',
    'untitled', 'copy of', 'kopie von', '(1)', '(2)', '(3)',
    'screenshot', 'bildschirmfoto', 'screen shot',
    'download', 'downloads'
]

# File extensions that are often temporary or unnecessary
TEMP_EXTENSIONS = [
    '.tmp', '.temp', '.bak', '.old', '.cache',
    '.crdownload', '.part', '.download'
]

# Archive/backup patterns
BACKUP_PATTERNS = [
    'backup', 'archive', 'archiv', 'old', 'alt',
    'sicherung', '.zip', '.tar', '.gz', '.7z', '.rar'
]

# Age thresholds (in days)
VERY_OLD_DAYS = 730  # 2 years
OLD_DAYS = 365  # 1 year
SOMEWHAT_OLD_DAYS = 180  # 6 months

# Size thresholds (in bytes)
LARGE_FILE_SIZE = 100 * 1024 * 1024  # 100 MB
VERY_LARGE_FILE_SIZE = 500 * 1024 * 1024  # 500 MB

# ============================================================================
# AUTHENTICATION
# ============================================================================

def authenticate(write_access=False):
    """Authenticate with Google Drive API."""
    creds = None
    scopes = SCOPES_WRITE if write_access else SCOPES_READONLY
    token_file = 'token_write.pickle' if write_access else 'token.pickle'

    if os.path.exists(token_file):
        with open(token_file, 'rb') as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists('credentials.json'):
                logger.error("credentials.json not found!")
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', scopes)
            creds = flow.run_local_server(port=0)

        with open(token_file, 'wb') as token:
            pickle.dump(creds, token)

    return build('drive', 'v3', credentials=creds)

# ============================================================================
# CONTENT EXTRACTION
# ============================================================================

class ContentExtractor:
    """Extract text content from various file types."""

    def __init__(self, service):
        self.service = service
        self.temp_dir = tempfile.mkdtemp(prefix='drive_cleanup_')
        logger.debug(f"Created temp directory: {self.temp_dir}")

    def cleanup(self):
        """Clean up temporary files."""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
            logger.debug(f"Cleaned up temp directory: {self.temp_dir}")

    def extract_text(self, file_id, mime_type, file_name):
        """Extract text content from a file."""
        try:
            # Google Docs types - export as plain text
            if mime_type == 'application/vnd.google-apps.document':
                return self._extract_google_doc(file_id)
            elif mime_type == 'application/vnd.google-apps.spreadsheet':
                return self._extract_google_sheet(file_id)
            elif mime_type == 'application/vnd.google-apps.presentation':
                return self._extract_google_slides(file_id)

            # PDF files
            elif mime_type == 'application/pdf' and HAS_PDF:
                return self._extract_pdf(file_id, file_name)

            # Word documents
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                              'application/msword'] and HAS_DOCX:
                return self._extract_word(file_id, file_name)

            # Excel files
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              'application/vnd.ms-excel'] and HAS_EXCEL:
                return self._extract_excel(file_id, file_name)

            # Plain text files
            elif mime_type.startswith('text/'):
                return self._extract_text_file(file_id)

            else:
                return None

        except Exception as e:
            logger.debug(f"Error extracting content from {file_name}: {e}")
            return None

    def _extract_google_doc(self, file_id):
        """Extract text from Google Doc."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/plain'
            )
            content = request.execute()
            return content.decode('utf-8', errors='ignore')
        except Exception as e:
            logger.debug(f"Error exporting Google Doc: {e}")
            return None

    def _extract_google_sheet(self, file_id):
        """Extract text from Google Sheet."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/csv'
            )
            content = request.execute()
            text = content.decode('utf-8', errors='ignore')
            # Return first few rows
            lines = text.split('\n')[:10]
            return '\n'.join(lines)
        except Exception as e:
            logger.debug(f"Error exporting Google Sheet: {e}")
            return None

    def _extract_google_slides(self, file_id):
        """Extract text from Google Slides."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/plain'
            )
            content = request.execute()
            return content.decode('utf-8', errors='ignore')
        except Exception as e:
            logger.debug(f"Error exporting Google Slides: {e}")
            return None

    def _download_file(self, file_id, file_name):
        """Download a file to temp directory."""
        try:
            request = self.service.files().get_media(fileId=file_id)
            file_path = os.path.join(self.temp_dir, file_name)

            with io.FileIO(file_path, 'wb') as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()

            return file_path
        except Exception as e:
            logger.debug(f"Error downloading file {file_name}: {e}")
            return None

    def _extract_pdf(self, file_id, file_name):
        """Extract text from PDF file."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            text = []
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                # Extract first 3 pages
                for i in range(min(3, len(pdf_reader.pages))):
                    page = pdf_reader.pages[i]
                    text.append(page.extract_text())

            os.remove(file_path)
            return '\n'.join(text)
        except Exception as e:
            logger.debug(f"Error reading PDF {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_word(self, file_id, file_name):
        """Extract text from Word document."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            doc = Document(file_path)
            # Extract first 20 paragraphs
            paragraphs = [p.text for p in doc.paragraphs[:20]]
            text = '\n'.join(paragraphs)

            os.remove(file_path)
            return text
        except Exception as e:
            logger.debug(f"Error reading Word doc {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_excel(self, file_id, file_name):
        """Extract text from Excel file."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            text = []

            # Get first sheet
            sheet = workbook.active
            # Get first 10 rows
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= 10:
                    break
                row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                text.append(row_text)

            workbook.close()
            os.remove(file_path)
            return '\n'.join(text)
        except Exception as e:
            logger.debug(f"Error reading Excel file {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_text_file(self, file_id):
        """Extract content from plain text file."""
        try:
            request = self.service.files().get_media(fileId=file_id)
            content = request.execute()
            # Return first 5000 characters
            return content.decode('utf-8', errors='ignore')[:5000]
        except Exception as e:
            logger.debug(f"Error reading text file: {e}")
            return None

    @staticmethod
    def create_summary(text, max_words=50):
        """Create a short summary from text (fallback method)."""
        if not text:
            return "Unable to extract content"

        # Clean up text
        text = re.sub(r'\s+', ' ', text).strip()

        # Split into words
        words = text.split()

        if len(words) <= max_words:
            return text

        # Return first max_words
        summary = ' '.join(words[:max_words]) + '...'
        return summary

    @staticmethod
    def create_claude_summary(text, file_name, bedrock_client):
        """Create an intelligent summary using Claude via AWS Bedrock."""
        if not text or not bedrock_client:
            return ContentExtractor.create_summary(text)

        # Truncate text if too long (Claude has token limits)
        max_chars = 15000
        if len(text) > max_chars:
            text = text[:max_chars] + "... [truncated]"

        prompt = f"""Analyze this file content and provide:
1. A brief summary (2-3 sentences) of what this file contains
2. An assessment of whether this file seems important or can likely be deleted

File name: {file_name}

Content:
{text}

Please respond in this format:
Summary: [your summary]
Assessment: [KEEP/DELETE] - [reasoning]"""

        try:
            # Prepare request body for Bedrock
            request_body = {
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 300,
                "messages": [{
                    "role": "user",
                    "content": prompt
                }]
            }

            # Call Bedrock
            response = bedrock_client.invoke_model(
                modelId="anthropic.claude-3-5-sonnet-20241022-v2:0",
                contentType="application/json",
                accept="application/json",
                body=json.dumps(request_body)
            )

            # Parse response
            response_body = json.loads(response['body'].read())
            return response_body['content'][0]['text'].strip()

        except Exception as e:
            logger.debug(f"Error using Claude via Bedrock: {e}")
            # Fallback to simple summary
            return ContentExtractor.create_summary(text)

# ============================================================================
# FILE ANALYSIS
# ============================================================================

class FileAnalyzer:
    def __init__(self, service, analyze_content=False, use_claude=False, aws_profile='dev', aws_region='us-east-1'):
        self.service = service
        self.analyze_content = analyze_content
        self.use_claude = use_claude
        self.content_extractor = ContentExtractor(service) if analyze_content else None
        self.bedrock_client = None

        # Initialize Bedrock client if requested
        if use_claude and HAS_BEDROCK:
            try:
                # Create boto3 session with specified profile
                session = boto3.Session(profile_name=aws_profile)
                self.bedrock_client = session.client(
                    service_name='bedrock-runtime',
                    region_name=aws_region
                )
                logger.info(f"AWS Bedrock initialized with profile '{aws_profile}' in region '{aws_region}' for Claude summaries")
            except Exception as e:
                logger.warning(f"Failed to initialize AWS Bedrock: {e}")
                logger.warning("Using fallback summary method")
                self.use_claude = False

        self.all_files = []
        self.all_folders = []
        self.delete_candidates = {
            'HIGH': [],  # Very confident to delete
            'MEDIUM': [],  # Probably safe to delete
            'LOW': []  # Consider deleting
        }
        self.stats = {
            'total_files': 0,
            'total_folders': 0,
            'total_size': 0,
            'potential_savings': 0,
            'content_analyzed': 0
        }

    def __del__(self):
        """Cleanup when analyzer is destroyed."""
        if self.content_extractor:
            self.content_extractor.cleanup()

    def analyze_filename(self, name):
        """Check if filename suggests it's temporary or obsolete."""
        name_lower = name.lower()
        reasons = []

        # Check for temp patterns
        for pattern in TEMP_PATTERNS:
            if pattern in name_lower:
                reasons.append(f"Name contains '{pattern}'")

        # Check for temp extensions
        for ext in TEMP_EXTENSIONS:
            if name_lower.endswith(ext):
                reasons.append(f"Temporary extension '{ext}'")

        # Check for backup patterns
        for pattern in BACKUP_PATTERNS:
            if pattern in name_lower:
                reasons.append(f"Backup/archive pattern '{pattern}'")

        return reasons

    def analyze_age(self, modified_time, viewed_time=None):
        """Analyze if file is old and unused."""
        reasons = []
        now = datetime.now(timezone.utc)

        modified = datetime.fromisoformat(modified_time.replace('Z', '+00:00'))
        age_days = (now - modified).days

        if age_days > VERY_OLD_DAYS:
            reasons.append(f"Very old: {age_days} days ({age_days//365} years) since modification")
        elif age_days > OLD_DAYS:
            reasons.append(f"Old: {age_days} days ({age_days//365} year) since modification")
        elif age_days > SOMEWHAT_OLD_DAYS:
            reasons.append(f"Somewhat old: {age_days} days ({age_days//30} months) since modification")

        # Check view time if available
        if viewed_time:
            viewed = datetime.fromisoformat(viewed_time.replace('Z', '+00:00'))
            view_age_days = (now - viewed).days
            if view_age_days > OLD_DAYS:
                reasons.append(f"Not viewed in {view_age_days} days ({view_age_days//365} years)")

        return reasons, age_days

    def analyze_size(self, size):
        """Analyze file size."""
        reasons = []

        if size == 0:
            reasons.append("Empty file (0 bytes)")
        elif size < 100:
            reasons.append(f"Very small file ({size} bytes)")
        elif size > VERY_LARGE_FILE_SIZE:
            reasons.append(f"Very large file ({size / (1024*1024):.1f} MB)")
        elif size > LARGE_FILE_SIZE:
            reasons.append(f"Large file ({size / (1024*1024):.1f} MB)")

        return reasons

    def classify_delete_confidence(self, reasons, age_days, size, mime_type=''):
        """Classify how confident we are this should be deleted."""

        # NEVER suggest deleting media files (photos, videos, audio)
        media_mime_types = [
            'image/', 'video/', 'audio/',
            'application/vnd.google-apps.photo',
            'application/vnd.google-apps.video'
        ]

        if any(mime_type.startswith(prefix) for prefix in media_mime_types):
            return None  # Never suggest deleting media files

        score = 0

        # Scoring based on reasons
        temp_keywords = ['tmp', 'temp', 'cache', 'untitled', 'screenshot', 'bildschirmfoto']
        backup_keywords = ['backup', 'archive', 'old', 'copy of', 'kopie von']

        reason_text = ' '.join(reasons).lower()

        # HIGH confidence indicators
        if any(k in reason_text for k in temp_keywords):
            score += 3
        if 'empty file' in reason_text:
            score += 3
        if size < 100 and age_days > 180:
            score += 2

        # MEDIUM confidence indicators
        if any(k in reason_text for k in backup_keywords):
            score += 2
        if age_days > VERY_OLD_DAYS:
            score += 2
        if 'not viewed' in reason_text:
            score += 1

        # LOW confidence indicators
        if age_days > OLD_DAYS:
            score += 1
        if size > VERY_LARGE_FILE_SIZE and age_days > SOMEWHAT_OLD_DAYS:
            score += 1

        # Classify
        if score >= 5:
            return 'HIGH'
        elif score >= 3:
            return 'MEDIUM'
        elif score >= 1:
            return 'LOW'
        else:
            return None

    def scan_folder(self, folder_id, max_files=10000):
        """Scan a specific folder and its subfolders recursively."""
        logger.info(f"Starting folder scan: {folder_id}")
        logger.info(f"Maximum files to scan: {max_files}")

        # Get folder info
        try:
            folder_info = self.service.files().get(
                fileId=folder_id,
                fields="id, name, mimeType"
            ).execute()
            logger.info(f"Scanning folder: {folder_info.get('name', 'Unknown')}")
        except Exception as e:
            logger.error(f"Error getting folder info: {e}")
            return

        # List to track folders to scan
        folders_to_scan = [folder_id]
        scanned_count = 0

        while folders_to_scan and scanned_count < max_files:
            current_folder = folders_to_scan.pop(0)

            # Get all items in current folder
            page_token = None

            while True:
                try:
                    results = self.service.files().list(
                        pageSize=1000,
                        pageToken=page_token,
                        fields="nextPageToken, files(id, name, mimeType, size, modifiedTime, viewedByMeTime, parents, trashed, webViewLink)",
                        q=f"'{current_folder}' in parents and trashed=false"
                    ).execute()

                    items = results.get('files', [])

                    for item in items:
                        scanned_count += 1

                        if scanned_count % 100 == 0:
                            logger.info(f"Scanned {scanned_count} items...")

                        # Separate files and folders
                        if item['mimeType'] == 'application/vnd.google-apps.folder':
                            self.all_folders.append(item)
                            self.stats['total_folders'] += 1
                            # Add subfolder to queue
                            folders_to_scan.append(item['id'])
                        else:
                            self.all_files.append(item)
                            self.stats['total_files'] += 1

                            # Add to total size (if available)
                            if 'size' in item:
                                self.stats['total_size'] += int(item['size'])

                        if scanned_count >= max_files:
                            logger.warning(f"Reached maximum scan limit of {max_files} files")
                            break

                    page_token = results.get('nextPageToken')
                    if not page_token or scanned_count >= max_files:
                        break

                except Exception as e:
                    logger.error(f"Error scanning folder {current_folder}: {e}")
                    break

            if scanned_count >= max_files:
                break

        logger.info(f"Scan complete: {self.stats['total_files']} files, {self.stats['total_folders']} folders")
        logger.info(f"Total size: {self.stats['total_size'] / (1024**3):.2f} GB")

    def scan_drive(self, max_files=10000):
        """Scan entire Google Drive."""
        logger.info("Starting Google Drive scan...")
        logger.info(f"Maximum files to scan: {max_files}")

        # Get all files and folders
        page_token = None
        scanned = 0

        while True:
            try:
                results = self.service.files().list(
                    pageSize=1000,
                    pageToken=page_token,
                    fields="nextPageToken, files(id, name, mimeType, size, modifiedTime, viewedByMeTime, parents, trashed, webViewLink)",
                    q="trashed=false"
                ).execute()

                items = results.get('files', [])

                for item in items:
                    scanned += 1

                    if scanned % 100 == 0:
                        logger.info(f"Scanned {scanned} items...")

                    # Separate files and folders
                    if item['mimeType'] == 'application/vnd.google-apps.folder':
                        self.all_folders.append(item)
                        self.stats['total_folders'] += 1
                    else:
                        self.all_files.append(item)
                        self.stats['total_files'] += 1

                        # Add to total size (if available)
                        if 'size' in item:
                            self.stats['total_size'] += int(item['size'])

                    if scanned >= max_files:
                        logger.warning(f"Reached maximum scan limit of {max_files} files")
                        break

                page_token = results.get('nextPageToken')
                if not page_token or scanned >= max_files:
                    break

            except Exception as e:
                logger.error(f"Error scanning drive: {e}")
                break

        logger.info(f"Scan complete: {self.stats['total_files']} files, {self.stats['total_folders']} folders")
        logger.info(f"Total size: {self.stats['total_size'] / (1024**3):.2f} GB")

    def analyze_files(self):
        """Analyze all files for deletion candidates."""
        logger.info("Analyzing files for deletion candidates...")

        for i, file_item in enumerate(self.all_files):
            if (i + 1) % 500 == 0:
                logger.info(f"Analyzed {i + 1}/{self.stats['total_files']} files...")

            name = file_item.get('name', 'Unknown')
            size = int(file_item.get('size', 0)) if 'size' in file_item else 0
            modified_time = file_item.get('modifiedTime')
            viewed_time = file_item.get('viewedByMeTime')
            mime_type = file_item.get('mimeType', 'Unknown')

            # Collect all reasons
            reasons = []

            # Analyze filename
            name_reasons = self.analyze_filename(name)
            reasons.extend(name_reasons)

            # Analyze age
            age_reasons, age_days = self.analyze_age(modified_time, viewed_time)
            reasons.extend(age_reasons)

            # Analyze size
            size_reasons = self.analyze_size(size)
            reasons.extend(size_reasons)

            # If we have reasons, classify confidence
            if reasons:
                confidence = self.classify_delete_confidence(reasons, age_days, size, mime_type)

                if confidence:
                    candidate = {
                        'id': file_item['id'],
                        'name': name,
                        'size': size,
                        'modified': modified_time,
                        'viewed': viewed_time,
                        'reasons': reasons,
                        'link': file_item.get('webViewLink', 'N/A'),
                        'age_days': age_days,
                        'mime_type': mime_type,
                        'summary': None
                    }

                    self.delete_candidates[confidence].append(candidate)
                    self.stats['potential_savings'] += size

        logger.info(f"Analysis complete!")
        logger.info(f"Found {len(self.delete_candidates['HIGH'])} HIGH confidence candidates")
        logger.info(f"Found {len(self.delete_candidates['MEDIUM'])} MEDIUM confidence candidates")
        logger.info(f"Found {len(self.delete_candidates['LOW'])} LOW confidence candidates")
        logger.info(f"Potential space savings: {self.stats['potential_savings'] / (1024**3):.2f} GB")

        # Analyze content if requested
        if self.analyze_content:
            self.analyze_content_for_candidates()

    def analyze_content_for_candidates(self):
        """Analyze content of delete candidates."""
        logger.info("")
        logger.info("Analyzing file content for delete candidates...")

        # Collect all candidates
        all_candidates = []
        for confidence in ['HIGH', 'MEDIUM', 'LOW']:
            all_candidates.extend(self.delete_candidates[confidence])

        # Filter candidates that can have content extracted
        extractable_mimes = [
            'application/vnd.google-apps.document',
            'application/vnd.google-apps.spreadsheet',
            'application/vnd.google-apps.presentation',
            'application/pdf',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/msword',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        ]

        extractable = [c for c in all_candidates if c['mime_type'] in extractable_mimes or c['mime_type'].startswith('text/')]

        logger.info(f"Found {len(extractable)} candidates with extractable content")
        logger.info(f"Analyzing content (this may take a while)...")

        # Limit files based on whether we're using Claude (more expensive) or not
        max_files = 50 if self.use_claude else 100

        for i, candidate in enumerate(extractable[:max_files]):
            if (i + 1) % 10 == 0:
                logger.info(f"  Extracted content from {i + 1}/{min(len(extractable), max_files)} files...")

            try:
                text = self.content_extractor.extract_text(
                    candidate['id'],
                    candidate['mime_type'],
                    candidate['name']
                )

                if text:
                    # Use Claude for intelligent summary if available
                    if self.use_claude and self.bedrock_client:
                        summary = ContentExtractor.create_claude_summary(
                            text,
                            candidate['name'],
                            self.bedrock_client
                        )
                    else:
                        summary = ContentExtractor.create_summary(text, max_words=50)

                    candidate['summary'] = summary
                    self.stats['content_analyzed'] += 1
            except Exception as e:
                logger.debug(f"Error analyzing content for {candidate['name']}: {e}")

        logger.info(f"Content analysis complete! Analyzed {self.stats['content_analyzed']} files")

    def analyze_empty_folders(self):
        """Find empty folders."""
        logger.info("Checking for empty folders...")

        empty_folders = []

        for folder in self.all_folders:
            folder_id = folder['id']
            folder_name = folder['name']

            # Check if folder has any children
            try:
                results = self.service.files().list(
                    pageSize=1,
                    q=f"'{folder_id}' in parents and trashed=false",
                    fields="files(id)"
                ).execute()

                files = results.get('files', [])

                if not files:
                    empty_folders.append({
                        'id': folder_id,
                        'name': folder_name,
                        'link': folder.get('webViewLink', 'N/A'),
                        'reasons': ['Empty folder (no files)']
                    })
            except Exception as e:
                logger.debug(f"Error checking folder {folder_name}: {e}")

        if empty_folders:
            self.delete_candidates['MEDIUM'].extend(empty_folders)
            logger.info(f"Found {len(empty_folders)} empty folders")

    def generate_report(self):
        """Generate a detailed report of deletion candidates."""
        report_lines = []
        report_lines.append("=" * 80)
        report_lines.append("GOOGLE DRIVE CLEANUP ANALYSIS REPORT")
        report_lines.append("=" * 80)
        report_lines.append("")

        # Summary
        report_lines.append("SUMMARY")
        report_lines.append("-" * 80)
        report_lines.append(f"Total files scanned: {self.stats['total_files']}")
        report_lines.append(f"Total folders scanned: {self.stats['total_folders']}")
        report_lines.append(f"Total size: {self.stats['total_size'] / (1024**3):.2f} GB")
        if self.analyze_content:
            report_lines.append(f"Files with content analyzed: {self.stats['content_analyzed']}")
        report_lines.append("")
        report_lines.append(f"Delete candidates found:")
        report_lines.append(f"  HIGH confidence:   {len([c for c in self.delete_candidates['HIGH'] if 'id' in c])} items")
        report_lines.append(f"  MEDIUM confidence: {len([c for c in self.delete_candidates['MEDIUM'] if 'id' in c])} items")
        report_lines.append(f"  LOW confidence:    {len([c for c in self.delete_candidates['LOW'] if 'id' in c])} items")
        report_lines.append("")
        report_lines.append(f"Potential space savings: {self.stats['potential_savings'] / (1024**3):.2f} GB")
        report_lines.append("")

        # Detailed candidates
        for confidence in ['HIGH', 'MEDIUM', 'LOW']:
            candidates = self.delete_candidates[confidence]
            if not candidates:
                continue

            report_lines.append("")
            report_lines.append("=" * 80)
            report_lines.append(f"{confidence} CONFIDENCE DELETE CANDIDATES")
            report_lines.append("=" * 80)

            # Sort by size (largest first)
            sorted_candidates = sorted(candidates, key=lambda x: x.get('size', 0), reverse=True)

            for i, candidate in enumerate(sorted_candidates[:50], 1):  # Show top 50
                size = candidate.get('size', 0)
                size_str = f"{size / (1024**2):.2f} MB" if size > 1024*1024 else f"{size / 1024:.2f} KB"

                report_lines.append("")
                report_lines.append(f"[{i}] {candidate['name']}")
                report_lines.append(f"    Size: {size_str}")
                report_lines.append(f"    Link: {candidate['link']}")
                report_lines.append(f"    Reasons:")
                for reason in candidate['reasons']:
                    report_lines.append(f"      - {reason}")

                # Add content summary if available
                if candidate.get('summary'):
                    report_lines.append(f"    Content Summary:")
                    report_lines.append(f"      {candidate['summary']}")

            if len(sorted_candidates) > 50:
                report_lines.append("")
                report_lines.append(f"... and {len(sorted_candidates) - 50} more items")

        report_lines.append("")
        report_lines.append("=" * 80)
        report_lines.append("END OF REPORT")
        report_lines.append("=" * 80)

        return "\n".join(report_lines)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def extract_folder_id(url_or_id):
    """Extract folder ID from URL or return ID if already provided."""
    if not url_or_id:
        return None

    # If it's already just an ID (no slashes or special chars)
    if '/' not in url_or_id and 'drive.google.com' not in url_or_id:
        return url_or_id

    # Extract from URL patterns
    import re

    # Pattern 1: /folders/FOLDER_ID
    match = re.search(r'/folders/([a-zA-Z0-9_-]+)', url_or_id)
    if match:
        return match.group(1)

    # Pattern 2: id=FOLDER_ID
    match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', url_or_id)
    if match:
        return match.group(1)

    return url_or_id

def extract_file_id_from_link(link):
    """Extract file ID from Google Drive link."""
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', link)
    if match:
        return match.group(1)
    match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', link)
    if match:
        return match.group(1)
    return None

def find_latest_report(folder_id):
    """Find the most recent cleanup report for a given folder ID."""
    # Pattern for report files
    pattern = f"drive_cleanup_report_{folder_id}_*.txt"
    reports = glob.glob(pattern)

    if not reports:
        return None

    # Sort by modification time (newest first)
    reports.sort(key=os.path.getmtime, reverse=True)
    return reports[0]

def get_single_key():
    """Get a single key press without requiring Enter."""
    if HAS_TERMIOS:
        # Unix/Linux/macOS
        fd = sys.stdin.fileno()
        old_settings = termios.tcgetattr(fd)
        try:
            tty.setraw(fd)
            ch = sys.stdin.read(1)
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)
        return ch
    elif HAS_MSVCRT:
        # Windows
        ch = msvcrt.getch()
        return ch.decode('utf-8')
    else:
        # Fallback to standard input
        return input().strip().lower()

def log_deleted_file(folder_id, file_name, file_link, file_size):
    """Log a deleted file to the deleted files list."""
    log_file = f"{folder_id}_deleted_files.txt"
    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')

    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{timestamp} | {file_name} | {file_size} | {file_link}\n")

def log_skipped_file(folder_id, file_name, file_link, file_size):
    """Log a skipped file to the skipped files list."""
    log_file = f"{folder_id}_skipped_files.txt"
    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')

    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{timestamp} | {file_name} | {file_size} | {file_link}\n")

def load_processed_files(folder_id):
    """Load already deleted and skipped files."""
    deleted_files = set()
    skipped_files = set()

    # Load deleted files
    deleted_log = f"{folder_id}_deleted_files.txt"
    if os.path.exists(deleted_log):
        with open(deleted_log, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(' | ')
                if len(parts) >= 3:
                    file_link = parts[3] if len(parts) > 3 else parts[2]
                    file_id = extract_file_id_from_link(file_link)
                    if file_id:
                        deleted_files.add(file_id)

    # Load skipped files
    skipped_log = f"{folder_id}_skipped_files.txt"
    if os.path.exists(skipped_log):
        with open(skipped_log, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(' | ')
                if len(parts) >= 3:
                    file_link = parts[3] if len(parts) > 3 else parts[2]
                    file_id = extract_file_id_from_link(file_link)
                    if file_id:
                        skipped_files.add(file_id)

    return deleted_files, skipped_files

def parse_cleanup_report(report_file):
    """Parse cleanup report and extract file information."""
    if not os.path.exists(report_file):
        logger.error(f"Report file not found: {report_file}")
        return []

    with open(report_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # Parse entries
    entries = []
    current_entry = None

    # Split by confidence sections
    sections = re.split(r'={80}\n([A-Z]+) CONFIDENCE DELETE CANDIDATES\n={80}', content)

    for i in range(1, len(sections), 2):
        confidence = sections[i]
        section_content = sections[i + 1]

        # Parse individual entries
        entry_pattern = r'\[(\d+)\] (.+?)\n    Size: (.+?)\n    Link: (.+?)\n    Reasons:\n((?:      - .+\n)+)'

        for match in re.finditer(entry_pattern, section_content):
            index, name, size, link, reasons_block = match.groups()

            reasons = [r.strip('- ').strip() for r in reasons_block.strip().split('\n')]

            # Extract file ID from link
            file_id = extract_file_id_from_link(link)

            # Check if there's a content summary
            summary_pattern = r'    Content Summary:\n      (.+?)(?:\n\n|\n\[|$)'
            summary_match = re.search(summary_pattern, section_content[match.end():])
            summary = summary_match.group(1) if summary_match else None

            entry = {
                'index': int(index),
                'name': name,
                'size': size,
                'link': link,
                'file_id': file_id,
                'confidence': confidence,
                'reasons': reasons,
                'summary': summary
            }
            entries.append(entry)

    return entries

def interactive_cleanup(service, report_file, folder_id):
    """Interactive cleanup session based on report."""
    import webbrowser
    from googleapiclient.errors import HttpError

    logger.info("=" * 80)
    logger.info("INTERACTIVE CLEANUP MODE")
    logger.info("=" * 80)
    logger.info(f"Loading report: {report_file}")

    entries = parse_cleanup_report(report_file)

    if not entries:
        logger.error("No entries found in report or report format is invalid")
        return

    logger.info(f"Found {len(entries)} delete candidates")

    # Load already processed files
    deleted_files, skipped_files = load_processed_files(folder_id)

    if deleted_files:
        logger.info(f"Found {len(deleted_files)} already deleted files - will skip those")
    if skipped_files:
        logger.info(f"Found {len(skipped_files)} already skipped files - will skip those")

    logger.info("")

    # Log files
    deleted_log = f"{folder_id}_deleted_files.txt"
    skipped_log = f"{folder_id}_skipped_files.txt"

    logger.info(f"Deleted files will be logged to: {deleted_log}")
    logger.info(f"Skipped files will be logged to: {skipped_log}")
    logger.info("")

    deleted_count = 0
    skipped_count = 0
    already_processed_count = 0

    for i, entry in enumerate(entries):
        # Skip if already deleted or skipped
        if entry['file_id'] in deleted_files:
            logger.debug(f"Skipping already deleted file: {entry['name']}")
            already_processed_count += 1
            continue

        if entry['file_id'] in skipped_files:
            logger.debug(f"Skipping already skipped file: {entry['name']}")
            already_processed_count += 1
            continue

        logger.info("=" * 80)
        logger.info(f"File {i + 1}/{len(entries)} - {entry['confidence']} CONFIDENCE")
        logger.info("=" * 80)
        logger.info(f"Name: {entry['name']}")
        logger.info(f"Size: {entry['size']}")
        logger.info(f"Link: {entry['link']}")
        logger.info(f"Reasons:")
        for reason in entry['reasons']:
            logger.info(f"  - {reason}")

        if entry['summary']:
            logger.info(f"Content Summary:")
            logger.info(f"  {entry['summary']}")

        logger.info("")

        while True:
            logger.info("Choose action: (1) Delete | (2) Open in Browser | (3) Skip | (q) Quit: ")
            choice = get_single_key().lower()
            logger.info(choice)  # Echo the choice

            if choice == '1':
                # Delete file
                if not entry['file_id']:
                    logger.error("Cannot delete: File ID not found")
                    break

                try:
                    service.files().delete(fileId=entry['file_id']).execute()
                    logger.info(f"✅ Deleted: {entry['name']}")
                    log_deleted_file(folder_id, entry['name'], entry['link'], entry['size'])
                    deleted_files.add(entry['file_id'])
                    deleted_count += 1
                    break
                except HttpError as e:
                    if e.resp.status == 404:
                        # File already deleted or doesn't exist
                        logger.warning(f"⚠️  File not found (404) - treating as deleted: {entry['name']}")
                        log_deleted_file(folder_id, entry['name'], entry['link'], entry['size'])
                        deleted_files.add(entry['file_id'])
                        deleted_count += 1
                        break
                    else:
                        logger.error(f"❌ Failed to delete: {e}")
                        break
                except Exception as e:
                    logger.error(f"❌ Failed to delete: {e}")
                    break

            elif choice == '2':
                # Open in browser
                logger.info(f"🌐 Opening in browser: {entry['link']}")
                try:
                    webbrowser.open(entry['link'])
                    logger.info("File opened in browser. Choose action:")
                    logger.info("")
                except Exception as e:
                    logger.error(f"Failed to open browser: {e}")
                    break

            elif choice == '3':
                # Skip
                logger.info(f"⏭️  Skipped: {entry['name']}")
                log_skipped_file(folder_id, entry['name'], entry['link'], entry['size'])
                skipped_files.add(entry['file_id'])
                skipped_count += 1
                break

            elif choice == 'q':
                logger.info("")
                logger.info("=" * 80)
                logger.info("CLEANUP SESSION SUMMARY")
                logger.info("=" * 80)
                logger.info(f"Files deleted: {deleted_count}")
                logger.info(f"Files skipped: {skipped_count}")
                logger.info(f"Files already processed: {already_processed_count}")
                logger.info(f"Files remaining: {len(entries) - i - 1}")
                logger.info("")
                logger.info(f"Logs saved:")
                logger.info(f"  Deleted: {deleted_log}")
                logger.info(f"  Skipped: {skipped_log}")
                return

            else:
                logger.warning("Invalid choice. Please press 1, 2, 3, or q")

        logger.info("")

    logger.info("=" * 80)
    logger.info("CLEANUP SESSION COMPLETE")
    logger.info("=" * 80)
    logger.info(f"Files deleted: {deleted_count}")
    logger.info(f"Files skipped: {skipped_count}")
    logger.info(f"Files already processed: {already_processed_count}")
    logger.info("")
    logger.info(f"Logs saved:")
    logger.info(f"  Deleted: {deleted_log}")
    logger.info(f"  Skipped: {skipped_log}")

# ============================================================================
# MAIN
# ============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Analyze Google Drive folder for cleanup candidates',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # Full workflow: Analyze AND clean a folder (default with Claude)
  python analyze_drive_cleanup.py "https://drive.google.com/drive/u/0/folders/1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh"

  # Only analyze (no interactive cleanup)
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --analyze

  # Only cleanup (interactive mode with existing report)
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --clean

  # Analyze without Claude (faster, simpler summaries)
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --analyze --no-claude

  # Use different AWS profile and region
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --aws-profile prod --aws-region eu-central-1

  # Analyze entire Drive (no folder parameter)
  python analyze_drive_cleanup.py --analyze
        '''
    )

    parser.add_argument(
        'folder',
        nargs='?',
        default=None,
        help='Google Drive folder URL or ID (required for most operations)'
    )

    parser.add_argument(
        '--analyze',
        action='store_true',
        help='Run analysis only (without interactive cleanup)'
    )

    parser.add_argument(
        '--clean',
        action='store_true',
        help='Run interactive cleanup only (without analysis)'
    )

    parser.add_argument(
        '--max-files',
        type=int,
        default=10000,
        help='Maximum number of files to scan (default: 10000)'
    )

    parser.add_argument(
        '--no-claude',
        action='store_true',
        help='Disable Claude AI summaries (use simple text extraction instead)'
    )

    parser.add_argument(
        '--aws-profile',
        type=str,
        default='dev',
        help='AWS profile to use for Bedrock (default: dev)'
    )

    parser.add_argument(
        '--aws-region',
        type=str,
        default='us-east-1',
        help='AWS region for Bedrock (default: us-east-1)'
    )

    args = parser.parse_args()

    # Record start time for logging
    start_time = datetime.now(timezone.utc)

    logger.info("Google Drive Cleanup Analyzer")
    logger.info("=" * 80)

    # Extract folder ID if provided
    folder_id = None
    if args.folder:
        folder_id = extract_folder_id(args.folder)
        logger.info(f"Target folder: {folder_id}")
    else:
        logger.info("Target: Entire Google Drive")
        folder_id = "full_drive"  # Use default for logging

    # Set up file logging
    setup_file_logging(folder_id, start_time)

    # Determine which operations to run
    run_analyze = args.analyze
    run_clean = args.clean

    # If neither specified, run both
    if not run_analyze and not run_clean:
        run_analyze = True
        run_clean = True
        logger.info("Mode: Full workflow (Analyze + Clean)")
    elif run_analyze and run_clean:
        logger.info("Mode: Full workflow (Analyze + Clean)")
    elif run_analyze:
        logger.info("Mode: Analysis only")
    elif run_clean:
        logger.info("Mode: Interactive cleanup only")

    logger.info("")

    # ========================================================================
    # ANALYSIS PHASE
    # ========================================================================
    if run_analyze:
        logger.info("=" * 80)
        logger.info("PHASE 1: ANALYSIS")
        logger.info("=" * 80)

        # Authenticate (read-only for analysis)
        logger.info("Authenticating with Google Drive API...")
        service = authenticate(write_access=False)
        logger.info("Authentication successful!")

        # Check if Claude should be used (default: yes, unless --no-claude)
        use_claude = not args.no_claude

        if use_claude:
            logger.info("Content analysis enabled with Claude AI")
            if not HAS_BEDROCK:
                logger.error("boto3 package not installed. Install with: pip install boto3")
                logger.error("Or use --no-claude for simple summaries")
                sys.exit(1)
            logger.info(f"AWS profile: {args.aws_profile}")
            logger.info(f"AWS region: {args.aws_region}")
        else:
            logger.info("Content analysis enabled (simple mode, no Claude)")

        # Check extraction libraries
        if not HAS_PDF:
            logger.warning("PyPDF2 not installed - PDF analysis will be skipped")
        if not HAS_DOCX:
            logger.warning("python-docx not installed - Word document analysis will be skipped")
        if not HAS_EXCEL:
            logger.warning("openpyxl not installed - Excel analysis will be skipped")

        # Create analyzer
        analyzer = FileAnalyzer(
            service,
            analyze_content=True,  # Always analyze content
            use_claude=use_claude,
            aws_profile=args.aws_profile,
            aws_region=args.aws_region
        )

        # Scan drive (folder-specific or entire drive)
        if folder_id:
            analyzer.scan_folder(folder_id, max_files=args.max_files)
        else:
            analyzer.scan_drive(max_files=args.max_files)

        # Analyze files
        analyzer.analyze_files()

        # Analyze folders
        analyzer.analyze_empty_folders()

        # Generate report
        report = analyzer.generate_report()

        # Save report
        folder_suffix = f"_{folder_id}" if folder_id else "_full_drive"
        report_filename = f"drive_cleanup_report{folder_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(report_filename, 'w', encoding='utf-8') as f:
            f.write(report)

        logger.info("")
        logger.info(f"✅ Report saved to: {report_filename}")
        logger.info("")

    # ========================================================================
    # CLEANUP PHASE
    # ========================================================================
    if run_clean:
        logger.info("=" * 80)
        logger.info("PHASE 2: INTERACTIVE CLEANUP")
        logger.info("=" * 80)

        # Find the report file
        if not folder_id:
            logger.error("Folder ID/URL is required for cleanup mode")
            sys.exit(1)

        report_file = find_latest_report(folder_id)

        if not report_file:
            logger.error(f"No cleanup report found for folder {folder_id}")
            logger.error(f"Expected file pattern: drive_cleanup_report_{folder_id}_*.txt")
            logger.error("Please run analysis first (without --clean flag)")
            sys.exit(1)

        logger.info(f"Using report: {report_file}")
        logger.info("")

        # Authenticate with write access
        logger.info("Authenticating with Google Drive API (write access)...")
        service_write = authenticate(write_access=True)
        logger.info("Authentication successful!")
        logger.info("")

        # Run interactive cleanup
        interactive_cleanup(service_write, report_file, folder_id)

if __name__ == '__main__':
    main()
