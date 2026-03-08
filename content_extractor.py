"""
Content extraction module for DriveCleanup
Handles extraction of text content from various file types
"""

import os
import io
import re
import tempfile
import json
from googleapiclient.http import MediaIoBaseDownload
from loguru import logger

from config import (
    MAX_PDF_PAGES, MAX_WORD_PARAGRAPHS, MAX_EXCEL_ROWS,
    MAX_TEXT_CHARS, MAX_CLAUDE_CHARS, MAX_SUMMARY_WORDS,
    PROMPTS_DIR
)

# Content extraction libraries
try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


class ContentExtractor:
    """Extract text content from various file types."""

    def __init__(self, service):
        self.service = service
        self.temp_dir = tempfile.mkdtemp(prefix='drive_cleanup_')
        logger.debug(f"Created temp directory: {self.temp_dir}")

    def cleanup(self):
        """Clean up temporary files."""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
            logger.debug(f"Cleaned up temp directory: {self.temp_dir}")

    def extract_text(self, file_id, mime_type, file_name):
        """Extract text content from a file."""
        try:
            # Google Docs types - export as plain text
            if mime_type == 'application/vnd.google-apps.document':
                return self._extract_google_doc(file_id)
            elif mime_type == 'application/vnd.google-apps.spreadsheet':
                return self._extract_google_sheet(file_id)
            elif mime_type == 'application/vnd.google-apps.presentation':
                return self._extract_google_slides(file_id)

            # PDF files
            elif mime_type == 'application/pdf' and HAS_PDF:
                return self._extract_pdf(file_id, file_name)

            # Word documents
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                              'application/msword'] and HAS_DOCX:
                return self._extract_word(file_id, file_name)

            # Excel files
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              'application/vnd.ms-excel'] and HAS_EXCEL:
                return self._extract_excel(file_id, file_name)

            # Plain text files
            elif mime_type.startswith('text/'):
                return self._extract_text_file(file_id)

            else:
                return None

        except Exception as e:
            logger.debug(f"Error extracting content from {file_name}: {e}")
            return None

    def _extract_google_doc(self, file_id):
        """Extract text from Google Doc."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/plain'
            )
            content = request.execute()
            return content.decode('utf-8', errors='ignore')
        except Exception as e:
            logger.debug(f"Error exporting Google Doc: {e}")
            return None

    def _extract_google_sheet(self, file_id):
        """Extract text from Google Sheet."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/csv'
            )
            content = request.execute()
            text = content.decode('utf-8', errors='ignore')
            # Return first few rows
            lines = text.split('\n')[:10]
            return '\n'.join(lines)
        except Exception as e:
            logger.debug(f"Error exporting Google Sheet: {e}")
            return None

    def _extract_google_slides(self, file_id):
        """Extract text from Google Slides."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/plain'
            )
            content = request.execute()
            return content.decode('utf-8', errors='ignore')
        except Exception as e:
            logger.debug(f"Error exporting Google Slides: {e}")
            return None

    def _download_file(self, file_id, file_name):
        """Download a file to temp directory."""
        try:
            request = self.service.files().get_media(fileId=file_id)
            file_path = os.path.join(self.temp_dir, file_name)

            with io.FileIO(file_path, 'wb') as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()

            return file_path
        except Exception as e:
            logger.debug(f"Error downloading file {file_name}: {e}")
            return None

    def _extract_pdf(self, file_id, file_name):
        """Extract text from PDF file."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            text = []
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                # Extract first few pages
                for i in range(min(MAX_PDF_PAGES, len(pdf_reader.pages))):
                    page = pdf_reader.pages[i]
                    text.append(page.extract_text())

            os.remove(file_path)
            return '\n'.join(text)
        except Exception as e:
            logger.debug(f"Error reading PDF {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_word(self, file_id, file_name):
        """Extract text from Word document."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            doc = Document(file_path)
            # Extract first paragraphs
            paragraphs = [p.text for p in doc.paragraphs[:MAX_WORD_PARAGRAPHS]]
            text = '\n'.join(paragraphs)

            os.remove(file_path)
            return text
        except Exception as e:
            logger.debug(f"Error reading Word doc {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_excel(self, file_id, file_name):
        """Extract text from Excel file."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            text = []

            # Get first sheet
            sheet = workbook.active
            # Get first rows
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= MAX_EXCEL_ROWS:
                    break
                row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                text.append(row_text)

            workbook.close()
            os.remove(file_path)
            return '\n'.join(text)
        except Exception as e:
            logger.debug(f"Error reading Excel file {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_text_file(self, file_id):
        """Extract content from plain text file."""
        try:
            request = self.service.files().get_media(fileId=file_id)
            content = request.execute()
            # Return first characters
            return content.decode('utf-8', errors='ignore')[:MAX_TEXT_CHARS]
        except Exception as e:
            logger.debug(f"Error reading text file: {e}")
            return None

    @staticmethod
    def create_summary(text, max_words=MAX_SUMMARY_WORDS):
        """Create a short summary from text (fallback method)."""
        if not text:
            return "Unable to extract content"

        # Clean up text
        text = re.sub(r'\s+', ' ', text).strip()

        # Split into words
        words = text.split()

        if len(words) <= max_words:
            return text

        # Return first max_words
        summary = ' '.join(words[:max_words]) + '...'
        return summary

    @staticmethod
    def create_claude_summary(text, file_name, bedrock_client):
        """Create an intelligent summary using Claude via AWS Bedrock."""
        if not text or not bedrock_client:
            return ContentExtractor.create_summary(text)

        # Truncate text if too long (Claude has token limits)
        if len(text) > MAX_CLAUDE_CHARS:
            text = text[:MAX_CLAUDE_CHARS] + "... [truncated]"

        # Load prompt template from file
        prompt_file = os.path.join(PROMPTS_DIR, 'claude_file_analysis.txt')
        try:
            with open(prompt_file, 'r', encoding='utf-8') as f:
                prompt_template = f.read()
            prompt = prompt_template.format(file_name=file_name, text=text)
        except FileNotFoundError:
            logger.warning(f"Prompt template not found: {prompt_file}, using fallback")
            # Fallback prompt
            prompt = f"""Analyze this file content and provide:
1. A brief summary (2-3 sentences) of what this file contains
2. An assessment of whether this file seems important or can likely be deleted

File name: {file_name}

Content:
{text}

Please respond in this format:
Summary: [your summary]
Assessment: [KEEP/DELETE] - [reasoning]"""

        try:
            # Prepare request body for Bedrock
            request_body = {
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 300,
                "messages": [{
                    "role": "user",
                    "content": prompt
                }]
            }

            # Call Bedrock
            response = bedrock_client.invoke_model(
                modelId="anthropic.claude-3-5-sonnet-20241022-v2:0",
                contentType="application/json",
                accept="application/json",
                body=json.dumps(request_body)
            )

            # Parse response
            response_body = json.loads(response['body'].read())
            return response_body['content'][0]['text'].strip()

        except Exception as e:
            logger.debug(f"Error using Claude via Bedrock: {e}")
            # Fallback to simple summary
            return ContentExtractor.create_summary(text)
