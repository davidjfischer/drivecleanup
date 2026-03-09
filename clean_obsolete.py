#!/usr/bin/env python3
"""
Google Drive Cleanup Analyzer
Analyzes Google Drive content and suggests files/folders to delete.
"""

import os
import pickle
import sys
import io
import tempfile
import re
import glob
import json
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from loguru import logger

# Import shared cleanup module
from cleanup_core import interactive_cleanup

# For single-key input
try:
    import termios
    import tty
    HAS_TERMIOS = True
except ImportError:
    HAS_TERMIOS = False

try:
    import msvcrt
    HAS_MSVCRT = True
except ImportError:
    HAS_MSVCRT = False

# Content extraction libraries
try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# AWS Bedrock for Claude
try:
    import boto3
    HAS_BEDROCK = True
except ImportError:
    HAS_BEDROCK = False

# OAuth Scopes
SCOPES_READONLY = ['https://www.googleapis.com/auth/drive.readonly']
SCOPES_WRITE = ['https://www.googleapis.com/auth/drive']  # For deletion

# Directories
STATE_DIR = 'state'
REPORTS_DIR = 'reports'
LOGS_DIR = 'logs'

# UI Box formatting constants
BOX_WIDTH = 78
BOX_TOTAL_WIDTH = 80  # BOX_WIDTH + 2 for borders
BOX_MAX_TEXT_WIDTH = 74  # For content with bullet points
BOX_CONTENT_MAX_WIDTH = 72  # For indented content

# Content extraction limits
MAX_PDF_PAGES = 3
MAX_WORD_PARAGRAPHS = 20
MAX_EXCEL_ROWS = 10
MAX_TEXT_CHARS = 5000
MAX_CLAUDE_CHARS = 15000

# Content analysis limits
MAX_SUMMARY_WORDS = 50

# Report display limits
MAX_CANDIDATES_IN_REPORT = 50

# Create directories
os.makedirs(STATE_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)

# Configure logger (stdout only initially, file logging added later)
logger.remove()
logger.add(
    sys.stdout,
    format="<green>{time:HH:mm:ss}</green> | <level>{level: <8}</level> | <level>{message}</level>",
    level="INFO",
    colorize=True
)

def setup_file_logging(folder_id, start_time):
    """Set up file logging for the session."""
    timestamp = start_time.strftime('%Y%m%d_%H%M%S')
    log_filename = os.path.join(LOGS_DIR, f"{folder_id}_session_{timestamp}.log")

    logger.debug(f"Setting up file logging to {log_filename}")

    # Add file handler
    logger.add(
        log_filename,
        format="{time:YYYY-MM-DD HH:mm:ss.SSS UTC} | {level: <8} | {message}",
        level="DEBUG",  # Log everything to file
        backtrace=True,
        diagnose=True
    )

    logger.info(f"Session log file: {log_filename}")
    return log_filename

# ============================================================================
# ANALYSIS CRITERIA
# ============================================================================

# File patterns that are typically safe to delete
TEMP_PATTERNS = [
    'tmp', 'temp', 'cache', '.cache', '.tmp',
    'untitled', 'copy of', 'kopie von', '(1)', '(2)', '(3)',
    'screenshot', 'bildschirmfoto', 'screen shot',
    'download', 'downloads'
]

# File extensions that are often temporary or unnecessary
TEMP_EXTENSIONS = [
    '.tmp', '.temp', '.bak', '.old', '.cache',
    '.crdownload', '.part', '.download'
]

# Archive/backup patterns
BACKUP_PATTERNS = [
    'backup', 'archive', 'archiv', 'old', 'alt',
    'sicherung', '.zip', '.tar', '.gz', '.7z', '.rar'
]

# Age thresholds (in days)
VERY_OLD_DAYS = 730  # 2 years
OLD_DAYS = 365  # 1 year
SOMEWHAT_OLD_DAYS = 180  # 6 months

# Size thresholds (in bytes)
LARGE_FILE_SIZE = 100 * 1024 * 1024  # 100 MB
VERY_LARGE_FILE_SIZE = 500 * 1024 * 1024  # 500 MB

# ============================================================================
# AUTHENTICATION
# ============================================================================

def authenticate(write_access=False):
    """Authenticate with Google Drive API."""
    access_type = "write" if write_access else "read-only"
    logger.debug(f"Authenticating with Google Drive API ({access_type} access)")

    creds = None
    scopes = SCOPES_WRITE if write_access else SCOPES_READONLY
    token_file = 'token_write.pickle' if write_access else 'token.pickle'

    logger.debug(f"Using token file: {token_file}")

    if os.path.exists(token_file):
        logger.debug(f"Loading existing credentials from {token_file}")
        with open(token_file, 'rb') as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            logger.info("Refreshing expired credentials")
            creds.refresh(Request())
        else:
            if not os.path.exists('credentials.json'):
                logger.error("credentials.json not found!")
                logger.error("Please download OAuth credentials from Google Cloud Console")
                sys.exit(1)

            logger.info("Starting OAuth flow - your browser will open")
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', scopes)
            creds = flow.run_local_server(port=0)

        logger.debug(f"Saving credentials to {token_file}")
        with open(token_file, 'wb') as token:
            pickle.dump(creds, token)

    logger.debug("Building Google Drive service")
    return build('drive', 'v3', credentials=creds)

# ============================================================================
# CONTENT EXTRACTION
# ============================================================================

class ContentExtractor:
    """Extract text content from various file types."""

    def __init__(self, service):
        self.service = service
        self.temp_dir = tempfile.mkdtemp(prefix='drive_cleanup_')
        logger.debug(f"Created temp directory: {self.temp_dir}")

    def cleanup(self):
        """Clean up temporary files."""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
            logger.debug(f"Cleaned up temp directory: {self.temp_dir}")

    def extract_text(self, file_id, mime_type, file_name):
        """Extract text content from a file."""
        try:
            # Google Docs types - export as plain text
            if mime_type == 'application/vnd.google-apps.document':
                return self._extract_google_doc(file_id)
            elif mime_type == 'application/vnd.google-apps.spreadsheet':
                return self._extract_google_sheet(file_id)
            elif mime_type == 'application/vnd.google-apps.presentation':
                return self._extract_google_slides(file_id)

            # PDF files
            elif mime_type == 'application/pdf' and HAS_PDF:
                return self._extract_pdf(file_id, file_name)

            # Word documents
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                              'application/msword'] and HAS_DOCX:
                return self._extract_word(file_id, file_name)

            # Excel files
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              'application/vnd.ms-excel'] and HAS_EXCEL:
                return self._extract_excel(file_id, file_name)

            # Plain text files
            elif mime_type.startswith('text/'):
                return self._extract_text_file(file_id)

            else:
                return None

        except Exception as e:
            logger.debug(f"Error extracting content from {file_name}: {e}")
            return None

    def _extract_google_doc(self, file_id):
        """Extract text from Google Doc."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/plain'
            )
            content = request.execute()
            return content.decode('utf-8', errors='ignore')
        except Exception as e:
            logger.debug(f"Error exporting Google Doc: {e}")
            return None

    def _extract_google_sheet(self, file_id):
        """Extract text from Google Sheet."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/csv'
            )
            content = request.execute()
            text = content.decode('utf-8', errors='ignore')
            # Return first few rows
            lines = text.split('\n')[:10]
            return '\n'.join(lines)
        except Exception as e:
            logger.debug(f"Error exporting Google Sheet: {e}")
            return None

    def _extract_google_slides(self, file_id):
        """Extract text from Google Slides."""
        try:
            request = self.service.files().export_media(
                fileId=file_id,
                mimeType='text/plain'
            )
            content = request.execute()
            return content.decode('utf-8', errors='ignore')
        except Exception as e:
            logger.debug(f"Error exporting Google Slides: {e}")
            return None

    def _download_file(self, file_id, file_name):
        """Download a file to temp directory."""
        try:
            request = self.service.files().get_media(fileId=file_id)
            file_path = os.path.join(self.temp_dir, file_name)

            with io.FileIO(file_path, 'wb') as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()

            return file_path
        except Exception as e:
            logger.debug(f"Error downloading file {file_name}: {e}")
            return None

    def _extract_pdf(self, file_id, file_name):
        """Extract text from PDF file."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            text = []
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                # Extract first few pages
                for i in range(min(MAX_PDF_PAGES, len(pdf_reader.pages))):
                    page = pdf_reader.pages[i]
                    text.append(page.extract_text())

            os.remove(file_path)
            return '\n'.join(text)
        except Exception as e:
            logger.debug(f"Error reading PDF {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_word(self, file_id, file_name):
        """Extract text from Word document."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            doc = Document(file_path)
            # Extract first paragraphs
            paragraphs = [p.text for p in doc.paragraphs[:MAX_WORD_PARAGRAPHS]]
            text = '\n'.join(paragraphs)

            os.remove(file_path)
            return text
        except Exception as e:
            logger.debug(f"Error reading Word doc {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_excel(self, file_id, file_name):
        """Extract text from Excel file."""
        file_path = self._download_file(file_id, file_name)
        if not file_path:
            return None

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            text = []

            # Get first sheet
            sheet = workbook.active
            # Get first rows
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= MAX_EXCEL_ROWS:
                    break
                row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                text.append(row_text)

            workbook.close()
            os.remove(file_path)
            return '\n'.join(text)
        except Exception as e:
            logger.debug(f"Error reading Excel file {file_name}: {e}")
            if os.path.exists(file_path):
                os.remove(file_path)
            return None

    def _extract_text_file(self, file_id):
        """Extract content from plain text file."""
        try:
            request = self.service.files().get_media(fileId=file_id)
            content = request.execute()
            # Return first characters
            return content.decode('utf-8', errors='ignore')[:MAX_TEXT_CHARS]
        except Exception as e:
            logger.debug(f"Error reading text file: {e}")
            return None

    @staticmethod
    def create_summary(text, max_words=50):
        """Create a short summary from text (fallback method)."""
        if not text:
            return "Unable to extract content"

        # Clean up text
        text = re.sub(r'\s+', ' ', text).strip()

        # Split into words
        words = text.split()

        if len(words) <= max_words:
            return text

        # Return first max_words
        summary = ' '.join(words[:max_words]) + '...'
        return summary

    @staticmethod
    def create_claude_summary(text, file_name, bedrock_client):
        """Create an intelligent summary using Claude via AWS Bedrock.

        Returns:
            dict with keys: 'summary', 'assessment', 'confidence'
            Or None if Claude analysis fails
        """
        if not text or not bedrock_client:
            return None

        # Truncate text if too long (Claude has token limits)
        if len(text) > MAX_CLAUDE_CHARS:
            text = text[:MAX_CLAUDE_CHARS] + "... [truncated]"

        prompt = f"""Analyze this file content and provide a structured assessment for cleanup decisions.

File name: {file_name}

Content:
{text}

Please respond in this EXACT format:
Summary: [2-3 sentence summary of what this file contains]
Assessment: [KEEP/DELETE]
Confidence: [HIGH/MEDIUM/LOW]
Reasoning: [brief explanation of your assessment]

Guidelines:
- DELETE if: temporary, outdated, backup, test file, obsolete content
- KEEP if: important documents, active projects, valuable data
- HIGH confidence: clearly obsolete or clearly important
- MEDIUM confidence: likely can be deleted but review recommended
- LOW confidence: uncertain, needs human judgment"""

        try:
            # Prepare request body for Bedrock
            request_body = {
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 400,
                "messages": [{
                    "role": "user",
                    "content": prompt
                }]
            }

            # Call Bedrock
            response = bedrock_client.invoke_model(
                modelId="anthropic.claude-3-5-sonnet-20241022-v2:0",
                contentType="application/json",
                accept="application/json",
                body=json.dumps(request_body)
            )

            # Parse response
            response_body = json.loads(response['body'].read())
            full_text = response_body['content'][0]['text'].strip()

            # Parse the structured response
            result = {
                'summary': None,
                'assessment': None,
                'confidence': None,
                'reasoning': None
            }

            for line in full_text.split('\n'):
                line = line.strip()
                if line.startswith('Summary:'):
                    result['summary'] = line.replace('Summary:', '').strip()
                elif line.startswith('Assessment:'):
                    result['assessment'] = line.replace('Assessment:', '').strip().upper()
                elif line.startswith('Confidence:'):
                    result['confidence'] = line.replace('Confidence:', '').strip().upper()
                elif line.startswith('Reasoning:'):
                    result['reasoning'] = line.replace('Reasoning:', '').strip()

            # Validate we got the key fields
            if result['summary'] and result['assessment']:
                logger.debug(f"Claude analysis successful for {file_name}: {result['assessment']}/{result.get('confidence', 'N/A')}")
                return result
            else:
                logger.warning(f"Incomplete Claude response for {file_name}. Got: {result}")
                logger.debug(f"Full Claude response: {full_text}")
                return None

        except Exception as e:
            logger.debug(f"Error using Claude via Bedrock: {e}")
            return None

# ============================================================================
# FILE ANALYSIS
# ============================================================================

class FileAnalyzer:
    def __init__(self, service, analyze_content=False, use_claude=False, aws_profile='dev', aws_region='us-east-1', min_age_days=90, skipped_files=None):
        self.service = service
        self.analyze_content = analyze_content
        self.use_claude = use_claude
        self.min_age_days = min_age_days
        self.skipped_files = skipped_files or set()
        self.content_extractor = ContentExtractor(service) if analyze_content else None
        self.bedrock_client = None

        # Initialize Bedrock client if requested
        if use_claude and HAS_BEDROCK:
            try:
                # Create boto3 session with specified profile
                session = boto3.Session(profile_name=aws_profile)
                self.bedrock_client = session.client(
                    service_name='bedrock-runtime',
                    region_name=aws_region
                )
                logger.info(f"AWS Bedrock initialized with profile '{aws_profile}' in region '{aws_region}' for Claude summaries")
            except Exception as e:
                logger.warning(f"Failed to initialize AWS Bedrock: {e}")
                logger.warning("Using fallback summary method")
                self.use_claude = False

        self.all_files = []
        self.all_folders = []  # All folders found during scan
        self.scanned_folders = []  # Only folders from current scan (for empty folder detection)
        self.delete_candidates = {
            'HIGH': [],  # Very confident to delete
            'MEDIUM': [],  # Probably safe to delete
            'LOW': []  # Consider deleting
        }
        self.stats = {
            'total_files': 0,
            'total_folders': 0,
            'total_size': 0,
            'potential_savings': 0,
            'content_analyzed': 0
        }
        # Track folder IDs to names for building paths
        self.folder_id_to_name = {}
        # Track folder IDs to their parent IDs for path traversal
        self.folder_id_to_parents = {}

    def __del__(self):
        """Cleanup when analyzer is destroyed."""
        if self.content_extractor:
            self.content_extractor.cleanup()

    def analyze_filename(self, name):
        """Check if filename suggests it's temporary or obsolete."""
        name_lower = name.lower()
        reasons = []

        # Check for temp patterns
        for pattern in TEMP_PATTERNS:
            if pattern in name_lower:
                reasons.append(f"Name contains '{pattern}'")

        # Check for temp extensions
        for ext in TEMP_EXTENSIONS:
            if name_lower.endswith(ext):
                reasons.append(f"Temporary extension '{ext}'")

        # Check for backup patterns
        for pattern in BACKUP_PATTERNS:
            if pattern in name_lower:
                reasons.append(f"Backup/archive pattern '{pattern}'")

        return reasons

    def analyze_age(self, modified_time, viewed_time=None):
        """Analyze if file is old and unused."""
        reasons = []
        now = datetime.now(timezone.utc)

        modified = datetime.fromisoformat(modified_time.replace('Z', '+00:00'))
        age_days = (now - modified).days

        # Check against user-configured minimum age threshold
        if age_days > self.min_age_days:
            # Categorize by severity
            if age_days > VERY_OLD_DAYS:
                reasons.append(f"Very old: {age_days} days ({age_days//365} years) since modification")
            elif age_days > OLD_DAYS:
                reasons.append(f"Old: {age_days} days ({age_days//365} year) since modification")
            elif age_days > SOMEWHAT_OLD_DAYS:
                reasons.append(f"Somewhat old: {age_days} days ({age_days//30} months) since modification")
            else:
                # Between min_age_days and SOMEWHAT_OLD_DAYS
                reasons.append(f"Not recently modified: {age_days} days ({age_days//30} months) since modification")

        # Check view time if available
        if viewed_time:
            viewed = datetime.fromisoformat(viewed_time.replace('Z', '+00:00'))
            view_age_days = (now - viewed).days
            if view_age_days > max(OLD_DAYS, self.min_age_days):
                reasons.append(f"Not viewed in {view_age_days} days ({view_age_days//365} years)")

        return reasons, age_days

    def analyze_size(self, size):
        """Analyze file size."""
        reasons = []

        if size == 0:
            reasons.append("Empty file (0 bytes)")
        elif size < 100:
            reasons.append(f"Very small file ({size} bytes)")
        elif size > VERY_LARGE_FILE_SIZE:
            reasons.append(f"Very large file ({size / (1024*1024):.1f} MB)")
        elif size > LARGE_FILE_SIZE:
            reasons.append(f"Large file ({size / (1024*1024):.1f} MB)")

        return reasons

    def classify_delete_confidence(self, reasons, age_days, size, mime_type=''):
        """Classify how confident we are this should be deleted."""

        # NEVER suggest deleting media files (photos, videos, audio)
        media_mime_types = [
            'image/', 'video/', 'audio/',
            'application/vnd.google-apps.photo',
            'application/vnd.google-apps.video'
        ]

        if any(mime_type.startswith(prefix) for prefix in media_mime_types):
            return None  # Never suggest deleting media files

        score = 0

        # Scoring based on reasons
        temp_keywords = ['tmp', 'temp', 'cache', 'untitled', 'screenshot', 'bildschirmfoto']
        backup_keywords = ['backup', 'archive', 'old', 'copy of', 'kopie von']

        reason_text = ' '.join(reasons).lower()

        # HIGH confidence indicators
        if any(k in reason_text for k in temp_keywords):
            score += 3
        if 'empty file' in reason_text:
            score += 3
        if size < 100 and age_days > 180:
            score += 2

        # MEDIUM confidence indicators
        if any(k in reason_text for k in backup_keywords):
            score += 2
        if age_days > VERY_OLD_DAYS:
            score += 2
        if 'not viewed' in reason_text:
            score += 1

        # LOW confidence indicators
        if age_days > OLD_DAYS:
            score += 1
        if size > VERY_LARGE_FILE_SIZE and age_days > SOMEWHAT_OLD_DAYS:
            score += 1
        # If file meets minimum age threshold but hasn't scored yet, give it LOW confidence
        if age_days > self.min_age_days and score == 0:
            score += 1

        # Classify
        if score >= 5:
            return 'HIGH'
        elif score >= 3:
            return 'MEDIUM'
        elif score >= 1:
            return 'LOW'
        else:
            return None

    def scan_folder(self, folder_id, max_files=10000):
        """Scan a specific folder and its subfolders recursively."""
        logger.info(f"Starting folder scan: {folder_id}")
        logger.info(f"Maximum files to scan: {max_files}")
        logger.debug(f"Scan mode: Recursive (includes subfolders)")

        # Get folder info
        try:
            logger.debug(f"Fetching folder metadata for {folder_id}")
            folder_info = self.service.files().get(
                fileId=folder_id,
                fields="id, name, mimeType"
            ).execute()
            folder_name = folder_info.get('name', 'Unknown')
            logger.info(f"Scanning folder: {folder_name}")
            logger.debug(f"Folder type: {folder_info.get('mimeType', 'Unknown')}")
        except Exception as e:
            logger.error(f"Error getting folder info: {e}")
            logger.error("Please check that the folder ID is correct and you have access")
            return

        # List to track folders to scan
        folders_to_scan = [folder_id]
        scanned_count = 0

        logger.debug("Starting recursive folder scan")
        while folders_to_scan and scanned_count < max_files:
            current_folder = folders_to_scan.pop(0)
            logger.debug(f"Scanning folder ID: {current_folder} ({len(folders_to_scan)} remaining in queue)")

            # Get all items in current folder
            page_token = None

            while True:
                try:
                    results = self.service.files().list(
                        pageSize=1000,
                        pageToken=page_token,
                        fields="nextPageToken, files(id, name, mimeType, size, modifiedTime, viewedByMeTime, parents, trashed, webViewLink)",
                        q=f"'{current_folder}' in parents and trashed=false and 'me' in owners"
                    ).execute()

                    items = results.get('files', [])

                    for item in items:
                        scanned_count += 1

                        if scanned_count % 100 == 0:
                            logger.info(f"Scanned {scanned_count} items...")

                        # Separate files and folders
                        if item['mimeType'] == 'application/vnd.google-apps.folder':
                            self.all_folders.append(item)
                            self.scanned_folders.append(item)  # Track for empty folder detection
                            self.stats['total_folders'] += 1
                            # Track folder names for path building
                            self.folder_id_to_name[item['id']] = item['name']
                            # Track folder parent relationships
                            if 'parents' in item:
                                self.folder_id_to_parents[item['id']] = item['parents']
                            # Add subfolder to queue
                            folders_to_scan.append(item['id'])
                        else:
                            self.all_files.append(item)
                            self.stats['total_files'] += 1

                            # Add to total size (if available)
                            if 'size' in item:
                                self.stats['total_size'] += int(item['size'])


                        if scanned_count >= max_files:
                            logger.warning(f"Reached maximum scan limit of {max_files} files")
                            break

                    page_token = results.get('nextPageToken')
                    if not page_token or scanned_count >= max_files:
                        break

                except Exception as e:
                    logger.error(f"Error scanning folder {current_folder}: {e}")
                    break

            if scanned_count >= max_files:
                break

        logger.info(f"Scan complete: {self.stats['total_files']} files, {self.stats['total_folders']} folders")
        logger.info(f"Total size: {self.stats['total_size'] / (1024**3):.2f} GB")

    def scan_drive(self, max_files=10000):
        """Scan entire Google Drive."""
        logger.info("Starting Google Drive scan...")
        logger.info(f"Maximum files to scan: {max_files}")

        # Get all files and folders
        page_token = None
        scanned = 0

        while True:
            try:
                results = self.service.files().list(
                    pageSize=1000,
                    pageToken=page_token,
                    fields="nextPageToken, files(id, name, mimeType, size, modifiedTime, viewedByMeTime, parents, trashed, webViewLink)",
                    q="trashed=false and 'me' in owners"
                ).execute()

                items = results.get('files', [])

                for item in items:
                    scanned += 1

                    if scanned % 100 == 0:
                        logger.info(f"Scanned {scanned} items...")

                    # Separate files and folders
                    if item['mimeType'] == 'application/vnd.google-apps.folder':
                        self.all_folders.append(item)
                        self.scanned_folders.append(item)  # Track for empty folder detection
                        self.stats['total_folders'] += 1
                        # Track folder names for path building
                        self.folder_id_to_name[item['id']] = item['name']
                        # Track folder parent relationships
                        if 'parents' in item:
                            self.folder_id_to_parents[item['id']] = item['parents']
                    else:
                        self.all_files.append(item)
                        self.stats['total_files'] += 1

                        # Add to total size (if available)
                        if 'size' in item:
                            self.stats['total_size'] += int(item['size'])


                    if scanned >= max_files:
                        logger.warning(f"Reached maximum scan limit of {max_files} files")
                        break

                page_token = results.get('nextPageToken')
                if not page_token or scanned >= max_files:
                    break

            except Exception as e:
                logger.error(f"Error scanning drive: {e}")
                break

        logger.info(f"Scan complete: {self.stats['total_files']} files, {self.stats['total_folders']} folders")
        logger.info(f"Total size: {self.stats['total_size'] / (1024**3):.2f} GB")

    def get_file_path(self, file_item):
        """Build the full path to a file from its parent folders."""
        path_parts = [file_item['name']]
        parents = file_item.get('parents', [])

        # Walk up the parent chain
        visited = set()
        while parents and len(parents) > 0:
            parent_id = parents[0]

            # Avoid infinite loops
            if parent_id in visited:
                break
            visited.add(parent_id)

            # Get parent folder name
            if parent_id in self.folder_id_to_name:
                path_parts.insert(0, self.folder_id_to_name[parent_id])

            # Find parent's parents using cached mapping
            if parent_id in self.folder_id_to_parents:
                parents = self.folder_id_to_parents[parent_id]
            else:
                # Fallback: search in all_folders (for backwards compatibility)
                parents = []
                for folder in self.all_folders:
                    if folder['id'] == parent_id:
                        parents = folder.get('parents', [])
                        break

        return '/'.join(path_parts)

    def analyze_files(self):
        """Analyze all files for deletion candidates."""
        logger.info("Analyzing files for deletion candidates...")

        for i, file_item in enumerate(self.all_files):
            if (i + 1) % 500 == 0:
                logger.info(f"Analyzed {i + 1}/{self.stats['total_files']} files...")

            name = file_item.get('name', 'Unknown')
            size = int(file_item.get('size', 0)) if 'size' in file_item else 0
            modified_time = file_item.get('modifiedTime')
            viewed_time = file_item.get('viewedByMeTime')
            mime_type = file_item.get('mimeType', 'Unknown')

            # Collect all reasons
            reasons = []

            # Analyze filename
            name_reasons = self.analyze_filename(name)
            reasons.extend(name_reasons)

            # Analyze age
            age_reasons, age_days = self.analyze_age(modified_time, viewed_time)
            reasons.extend(age_reasons)

            # Analyze size
            size_reasons = self.analyze_size(size)
            reasons.extend(size_reasons)

            # If we have reasons, classify confidence
            if reasons:
                confidence = self.classify_delete_confidence(reasons, age_days, size, mime_type)

                if confidence:
                    candidate = {
                        'id': file_item['id'],
                        'name': name,
                        'size': size,
                        'modified': modified_time,
                        'viewed': viewed_time,
                        'reasons': reasons,
                        'link': file_item.get('webViewLink', 'N/A'),
                        'age_days': age_days,
                        'mime_type': mime_type,
                        'summary': None
                    }

                    self.delete_candidates[confidence].append(candidate)
                    self.stats['potential_savings'] += size

        logger.info(f"Analysis complete!")
        logger.info(f"Found {len(self.delete_candidates['HIGH'])} HIGH confidence candidates")
        logger.info(f"Found {len(self.delete_candidates['MEDIUM'])} MEDIUM confidence candidates")
        logger.info(f"Found {len(self.delete_candidates['LOW'])} LOW confidence candidates")
        logger.info(f"Potential space savings: {self.stats['potential_savings'] / (1024**3):.2f} GB")

        # Analyze content if requested
        if self.analyze_content:
            self.analyze_content_for_candidates()

    def analyze_content_for_candidates(self):
        """Analyze content of delete candidates."""
        logger.info("")
        logger.info("Analyzing file content for delete candidates...")

        # Collect all candidates
        all_candidates = []
        for confidence in ['HIGH', 'MEDIUM', 'LOW']:
            all_candidates.extend(self.delete_candidates[confidence])

        # Filter candidates that can have content extracted
        extractable_mimes = [
            'application/vnd.google-apps.document',
            'application/vnd.google-apps.spreadsheet',
            'application/vnd.google-apps.presentation',
            'application/pdf',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/msword',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        ]

        extractable = [c for c in all_candidates if c['mime_type'] in extractable_mimes or c['mime_type'].startswith('text/')]

        # Filter out skipped files
        filtered_extractable = []
        skipped_count = 0

        for candidate in extractable:
            # Skip if this file was already skipped in a previous run
            if candidate['id'] in self.skipped_files:
                skipped_count += 1
                continue

            filtered_extractable.append(candidate)

        if skipped_count > 0:
            logger.info(f"Skipping content analysis for {skipped_count} previously skipped files")

        logger.info(f"Found {len(filtered_extractable)} candidates requiring content analysis")
        if len(filtered_extractable) > 0:
            logger.info(f"Analyzing content (this may take a while)...")

        for i, candidate in enumerate(filtered_extractable):
            if (i + 1) % 10 == 0:
                logger.info(f"  Extracted content from {i + 1}/{len(filtered_extractable)} files...")

            try:
                text = self.content_extractor.extract_text(
                    candidate['id'],
                    candidate['mime_type'],
                    candidate['name']
                )

                if text:
                    # Use Claude for intelligent summary if available
                    if self.use_claude and self.bedrock_client:
                        logger.debug(f"Requesting Claude analysis for: {candidate['name']}")
                        claude_result = ContentExtractor.create_claude_summary(
                            text,
                            candidate['name'],
                            self.bedrock_client
                        )

                        if claude_result:
                            # Use Claude's summary
                            candidate['summary'] = claude_result['summary']
                            logger.debug(f"Using Claude summary for: {candidate['name']}")

                            # Use Claude's confidence if available and assessment is DELETE
                            if claude_result.get('assessment') == 'DELETE' and claude_result.get('confidence'):
                                # Store Claude's reasoning
                                if claude_result.get('reasoning'):
                                    candidate['reasons'].append(f"Claude assessment: {claude_result['reasoning']}")

                                # Adjust confidence based on Claude's assessment
                                claude_confidence = claude_result['confidence']
                                current_confidence = None

                                # Find current confidence level
                                for conf_level in ['HIGH', 'MEDIUM', 'LOW']:
                                    if candidate in self.delete_candidates[conf_level]:
                                        current_confidence = conf_level
                                        break

                                # Move to Claude's suggested confidence if different
                                if current_confidence and claude_confidence in ['HIGH', 'MEDIUM', 'LOW']:
                                    if current_confidence != claude_confidence:
                                        logger.debug(f"Claude adjusted confidence for {candidate['name']}: {current_confidence} → {claude_confidence}")
                                        self.delete_candidates[current_confidence].remove(candidate)
                                        self.delete_candidates[claude_confidence].append(candidate)

                            elif claude_result.get('assessment') == 'KEEP':
                                # Claude says KEEP - remove from delete candidates
                                logger.debug(f"Claude recommends KEEP for {candidate['name']}: {claude_result.get('reasoning')}")
                                for conf_level in ['HIGH', 'MEDIUM', 'LOW']:
                                    if candidate in self.delete_candidates[conf_level]:
                                        self.delete_candidates[conf_level].remove(candidate)
                                        logger.info(f"  Removed {candidate['name']} from candidates (Claude: KEEP)")
                                        break

                            self.stats['content_analyzed'] += 1
                        else:
                            # Claude failed, use fallback
                            logger.info(f"  Claude analysis failed for {candidate['name']}, using fallback summary")
                            summary = ContentExtractor.create_summary(text, max_words=MAX_SUMMARY_WORDS)
                            candidate['summary'] = summary
                            self.stats['content_analyzed'] += 1
                    else:
                        # No Claude, use simple summary
                        logger.debug(f"Claude not enabled, using simple summary for: {candidate['name']}")
                        summary = ContentExtractor.create_summary(text, max_words=MAX_SUMMARY_WORDS)
                        candidate['summary'] = summary
                        self.stats['content_analyzed'] += 1
            except Exception as e:
                logger.debug(f"Error analyzing content for {candidate['name']}: {e}")

        logger.info(f"Content analysis complete! Analyzed {self.stats['content_analyzed']} files")

    def _folder_contains_only_empty_folders(self, folder_id, checked_folders=None):
        """Recursively check if a folder contains only empty folders (no files)."""
        if checked_folders is None:
            checked_folders = set()

        # Avoid infinite loops
        if folder_id in checked_folders:
            return True
        checked_folders.add(folder_id)

        try:
            logger.debug(f"Checking if folder {folder_id} contains only empty folders")

            # Get all children
            results = self.service.files().list(
                pageSize=1000,
                q=f"'{folder_id}' in parents and trashed=false and 'me' in owners",
                fields="files(id, mimeType)"
            ).execute()

            children = results.get('files', [])

            if not children:
                # Completely empty
                logger.debug(f"Folder {folder_id} is completely empty")
                return True

            # Check if there are any files (non-folders)
            for child in children:
                if child['mimeType'] != 'application/vnd.google-apps.folder':
                    # Found a file - not empty
                    logger.debug(f"Folder {folder_id} contains files")
                    return False

            # Only folders - check recursively
            logger.debug(f"Folder {folder_id} contains only subfolders, checking recursively")
            for child in children:
                if not self._folder_contains_only_empty_folders(child['id'], checked_folders):
                    return False

            # All subfolders are empty
            logger.debug(f"Folder {folder_id} contains only empty subfolders")
            return True

        except Exception as e:
            logger.debug(f"Error checking folder {folder_id}: {e}")
            return False

    def analyze_empty_folders(self):
        """Find empty folders and folders containing only empty subfolders."""
        if not self.scanned_folders:
            logger.info("No folders found in scanned area, skipping empty folder analysis")
            return

        logger.info("Checking for empty folders and folders with only empty subfolders...")

        empty_folders = []
        folders_with_only_empty_subfolders = []

        for folder in self.scanned_folders:
            folder_id = folder['id']
            folder_name = folder['name']

            logger.debug(f"Analyzing folder: {folder_name}")

            # Check if folder contains only empty folders recursively
            try:
                if self._folder_contains_only_empty_folders(folder_id):
                    # Check if completely empty or has empty subfolders
                    results = self.service.files().list(
                        pageSize=1,
                        q=f"'{folder_id}' in parents and trashed=false and 'me' in owners",
                        fields="files(id)"
                    ).execute()

                    if not results.get('files', []):
                        # Completely empty
                        empty_folders.append({
                            'id': folder_id,
                            'name': folder_name,
                            'link': folder.get('webViewLink', 'N/A'),
                            'reasons': ['Empty folder (no files)'],
                            'size': 0
                        })
                        logger.debug(f"Marked as empty: {folder_name}")
                    else:
                        # Has subfolders, but all empty
                        folders_with_only_empty_subfolders.append({
                            'id': folder_id,
                            'name': folder_name,
                            'link': folder.get('webViewLink', 'N/A'),
                            'reasons': ['Folder contains only empty subfolders (no files)'],
                            'size': 0
                        })
                        logger.debug(f"Marked as containing only empty subfolders: {folder_name}")

            except Exception as e:
                logger.debug(f"Error checking folder {folder_name}: {e}")

        # Add to delete candidates
        if empty_folders:
            self.delete_candidates['HIGH'].extend(empty_folders)
            logger.info(f"Found {len(empty_folders)} completely empty folders")

        if folders_with_only_empty_subfolders:
            self.delete_candidates['HIGH'].extend(folders_with_only_empty_subfolders)
            logger.info(f"Found {len(folders_with_only_empty_subfolders)} folders containing only empty subfolders")

    def generate_report(self):
        """Generate a detailed JSON report of deletion candidates."""
        # Prepare candidates by confidence level
        candidates_by_confidence = {}
        for confidence in ['HIGH', 'MEDIUM', 'LOW']:
            candidates = self.delete_candidates[confidence]
            if candidates:
                # Sort by size (largest first)
                sorted_candidates = sorted(candidates, key=lambda x: x.get('size', 0), reverse=True)
                # Limit to MAX_CANDIDATES_IN_REPORT
                candidates_by_confidence[confidence] = sorted_candidates[:MAX_CANDIDATES_IN_REPORT]
                if len(sorted_candidates) > MAX_CANDIDATES_IN_REPORT:
                    candidates_by_confidence[f"{confidence}_total"] = len(sorted_candidates)

        # Build JSON report structure
        report_data = {
            "report_type": "Google Drive Cleanup Analysis",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {
                "total_files_scanned": self.stats['total_files'],
                "total_folders_scanned": self.stats['total_folders'],
                "total_size_bytes": self.stats['total_size'],
                "total_size_gb": round(self.stats['total_size'] / (1024**3), 2),
                "content_analyzed": self.stats['content_analyzed'] if self.analyze_content else 0,
                "delete_candidates": {
                    "HIGH": len([c for c in self.delete_candidates['HIGH'] if 'id' in c]),
                    "MEDIUM": len([c for c in self.delete_candidates['MEDIUM'] if 'id' in c]),
                    "LOW": len([c for c in self.delete_candidates['LOW'] if 'id' in c])
                },
                "potential_savings_bytes": self.stats['potential_savings'],
                "potential_savings_gb": round(self.stats['potential_savings'] / (1024**3), 2)
            },
            "candidates": candidates_by_confidence
        }

        return json.dumps(report_data, indent=2, ensure_ascii=False)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def extract_folder_id(url_or_id):
    """Extract folder ID from URL or return ID if already provided."""
    if not url_or_id:
        return None

    # If it's already just an ID (no slashes or special chars)
    if '/' not in url_or_id and 'drive.google.com' not in url_or_id:
        return url_or_id

    # Extract from URL patterns
    import re

    # Pattern 1: /folders/FOLDER_ID
    match = re.search(r'/folders/([a-zA-Z0-9_-]+)', url_or_id)
    if match:
        return match.group(1)

    # Pattern 2: id=FOLDER_ID
    match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', url_or_id)
    if match:
        return match.group(1)

    return url_or_id

def extract_file_id_from_link(link):
    """Extract file ID from Google Drive link."""
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', link)
    if match:
        return match.group(1)
    match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', link)
    if match:
        return match.group(1)
    return None

def find_latest_report(folder_id):
    """Find the most recent cleanup report for a given folder ID."""
    logger.debug(f"Searching for reports in {REPORTS_DIR} directory")

    # Pattern for report files
    pattern = os.path.join(REPORTS_DIR, f"drive_cleanup_report_{folder_id}_*.json")
    reports = glob.glob(pattern)

    logger.debug(f"Found {len(reports)} report(s) matching pattern")

    if not reports:
        logger.warning(f"No reports found for folder {folder_id}")
        return None

    # Sort by modification time (newest first)
    reports.sort(key=os.path.getmtime, reverse=True)
    latest = reports[0]
    logger.info(f"Using latest report: {os.path.basename(latest)}")
    return latest

def get_single_key():
    """Get a single key press without requiring Enter."""
    if HAS_TERMIOS:
        # Unix/Linux/macOS
        fd = sys.stdin.fileno()
        old_settings = termios.tcgetattr(fd)
        try:
            tty.setraw(fd)
            ch = sys.stdin.read(1)
        finally:
            termios.tcsetattr(fd, termios.TCSADRAIN, old_settings)
        return ch
    elif HAS_MSVCRT:
        # Windows
        ch = msvcrt.getch()
        return ch.decode('utf-8')
    else:
        # Fallback to standard input
        return input().strip().lower()

def log_deleted_file(folder_id, file_name, file_link, file_size):
    """Log a deleted file to the deleted files list."""
    log_file = os.path.join(STATE_DIR, f"{folder_id}_deleted_files.txt")
    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')

    logger.debug(f"Logging deleted file to {log_file}: {file_name}")

    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{timestamp} | {file_name} | {file_size} | {file_link}\n")

def log_skipped_file(folder_id, file_name, file_link, file_size):
    """Log a skipped file to the skipped files list."""
    log_file = os.path.join(STATE_DIR, f"{folder_id}_skipped_files.txt")
    timestamp = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')

    logger.debug(f"Logging skipped file to {log_file}: {file_name}")

    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{timestamp} | {file_name} | {file_size} | {file_link}\n")

def load_processed_files(folder_id):
    """Load already deleted and skipped files."""
    logger.debug(f"Loading processed files for folder {folder_id}")

    deleted_files = set()
    skipped_files = set()

    # Load deleted files
    deleted_log = os.path.join(STATE_DIR, f"{folder_id}_deleted_files.txt")
    if os.path.exists(deleted_log):
        logger.debug(f"Loading deleted files from {deleted_log}")
        with open(deleted_log, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(' | ')
                if len(parts) >= 3:
                    file_link = parts[3] if len(parts) > 3 else parts[2]
                    file_id = extract_file_id_from_link(file_link)
                    if file_id:
                        deleted_files.add(file_id)
        logger.debug(f"Loaded {len(deleted_files)} deleted file IDs")
    else:
        logger.debug(f"No deleted files log found at {deleted_log}")

    # Load skipped files
    skipped_log = os.path.join(STATE_DIR, f"{folder_id}_skipped_files.txt")
    if os.path.exists(skipped_log):
        logger.debug(f"Loading skipped files from {skipped_log}")
        with open(skipped_log, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(' | ')
                if len(parts) >= 3:
                    file_link = parts[3] if len(parts) > 3 else parts[2]
                    file_id = extract_file_id_from_link(file_link)
                    if file_id:
                        skipped_files.add(file_id)
        logger.debug(f"Loaded {len(skipped_files)} skipped file IDs")
    else:
        logger.debug(f"No skipped files log found at {skipped_log}")

    return deleted_files, skipped_files

def parse_cleanup_report(report_file):
    """Parse JSON cleanup report and extract file information."""
    if not os.path.exists(report_file):
        logger.error(f"Report file not found: {report_file}")
        return []

    try:
        with open(report_file, 'r', encoding='utf-8') as f:
            report_data = json.load(f)
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse JSON report: {e}")
        return []

    # Extract entries from all confidence levels
    entries = []
    candidates = report_data.get('candidates', {})

    for confidence in ['HIGH', 'MEDIUM', 'LOW']:
        if confidence in candidates:
            for candidate in candidates[confidence]:
                # Format size for display (similar to old format)
                size = candidate.get('size', 0)
                if size > 1024*1024:
                    size_str = f"{size / (1024**2):.2f} MB"
                else:
                    size_str = f"{size / 1024:.2f} KB"

                # Extract file ID from link
                file_id = extract_file_id_from_link(candidate.get('link', ''))

                entry = {
                    'index': len(entries) + 1,  # Sequential index
                    'name': candidate.get('name', 'Unknown'),
                    'path': candidate.get('path'),  # May be None
                    'size': size_str,
                    'link': candidate.get('link', 'N/A'),
                    'file_id': file_id,
                    'confidence': confidence,
                    'reasons': candidate.get('reasons', []),
                    'summary': candidate.get('summary')
                }
                entries.append(entry)

    return entries

def format_box_line(text, width=BOX_WIDTH, color_code=None):
    """Format a line for the dialog box with exact width."""
    # Account for "║ " and " ║" (4 characters total)
    max_text_width = width - 2

    # Calculate visible length (accounting for color codes and invisible Unicode)
    import re
    import unicodedata

    # First, remove ANSI color codes
    visible_text = re.sub(r'\x1b\[[0-9;]+m', '', text)

    # Remove invisible Unicode characters (BOM, zero-width spaces, etc.)
    # Keep the original text but calculate length based on visible characters
    cleaned_for_length = ''.join(
        char for char in visible_text
        if unicodedata.category(char) not in ['Cf', 'Cc']  # Format/Control characters
    )

    # Also strip BOM specifically
    cleaned_for_length = cleaned_for_length.replace('\ufeff', '')  # BOM
    cleaned_for_length = cleaned_for_length.replace('\u200b', '')  # Zero-width space

    visible_length = len(cleaned_for_length)

    # Ensure text doesn't exceed max width (truncate if needed)
    if visible_length > max_text_width:
        # Truncate to fit - use cleaned version
        excess = visible_length - max_text_width
        text = text[:len(text) - excess]
        visible_length = max_text_width

    # Pad with spaces to exact width based on visible length
    padding = max_text_width - visible_length
    line = "║ " + text + " " * padding + " ║"

    if color_code:
        reset = '\033[0m'
        return f"{color_code}{line}{reset}"
    return line

def format_box_separator(char="─", width=BOX_WIDTH, color_code=None):
    """Format a separator line for the dialog box."""
    line = "╠" + char * width + "╣"

    if color_code:
        reset = '\033[0m'
        return f"{color_code}{line}{reset}"
    return line

def get_display_width(text):
    """
    Calculate the display width of text, accounting for emojis being 2 characters wide.
    """
    import unicodedata
    width = 0
    for char in text:
        # Check if character is an emoji or wide character
        if unicodedata.east_asian_width(char) in ('F', 'W'):
            width += 2
        else:
            width += 1
    return width

def print_colored_tip_box(lines, color_code='\033[91m'):
    """
    Print a colored tip box with multiple lines of text.
    Default color is red (91m).
    Color codes: 91m=red, 92m=green, 93m=yellow, 94m=blue, 95m=magenta, 96m=cyan
    """
    reset = '\033[0m'
    width = BOX_WIDTH

    print()
    print(f"{color_code}╔" + "═" * width + "╗" + reset)

    for line in lines:
        # Calculate actual display width
        display_width = get_display_width(line)
        padding_needed = width - 2 - display_width

        # Word wrap if needed
        if display_width <= width - 2:
            print(f"{color_code}║ {line}{' ' * padding_needed} ║{reset}")
        else:
            # Word wrap considering display width
            words = line.split()
            current_line = ""
            current_width = 0

            for word in words:
                word_with_space = word + " "
                word_width = get_display_width(word_with_space)

                if current_width + word_width <= width - 2:
                    current_line += word_with_space
                    current_width += word_width
                else:
                    if current_line:
                        line_stripped = current_line.rstrip()
                        line_display_width = get_display_width(line_stripped)
                        padding = width - 2 - line_display_width
                        print(f"{color_code}║ {line_stripped}{' ' * padding} ║{reset}")
                    current_line = word + " "
                    current_width = get_display_width(current_line)

            if current_line:
                line_stripped = current_line.rstrip()
                line_display_width = get_display_width(line_stripped)
                padding = width - 2 - line_display_width
                print(f"{color_code}║ {line_stripped}{' ' * padding} ║{reset}")

    print(f"{color_code}╚" + "═" * width + "╝" + reset)
    print()

def interactive_cleanup(service, report_file, folder_id):
    """Interactive cleanup session based on report."""
    import webbrowser
    from googleapiclient.errors import HttpError

    logger.info("=" * 80)
    logger.info("INTERACTIVE CLEANUP MODE")
    logger.info("=" * 80)
    logger.info(f"Loading report: {report_file}")

    entries = parse_cleanup_report(report_file)

    if not entries:
        logger.error("No entries found in report or report format is invalid")
        return

    logger.info(f"Found {len(entries)} delete candidates")

    # Load already processed files
    deleted_files, skipped_files = load_processed_files(folder_id)

    if deleted_files:
        logger.info(f"Found {len(deleted_files)} already deleted files - will skip those")
    if skipped_files:
        logger.info(f"Found {len(skipped_files)} already skipped files - will skip those")

    logger.info("")

    # Log files
    deleted_log = os.path.join(STATE_DIR, f"{folder_id}_deleted_files.txt")
    skipped_log = os.path.join(STATE_DIR, f"{folder_id}_skipped_files.txt")

    logger.info(f"Deleted files will be logged to: {deleted_log}")
    logger.info(f"Skipped files will be logged to: {skipped_log}")
    logger.info("")

    deleted_count = 0
    skipped_count = 0
    already_processed_count = 0

    for i, entry in enumerate(entries):
        # Skip if already deleted or skipped
        if entry['file_id'] in deleted_files:
            logger.debug(f"Skipping already deleted file: {entry['name']}")
            already_processed_count += 1
            continue

        if entry['file_id'] in skipped_files:
            logger.debug(f"Skipping already skipped file: {entry['name']}")
            already_processed_count += 1
            continue

        # Log file description BEFORE showing dialog
        logger.info("=" * 80)
        logger.info(f"Presenting file {i + 1}/{len(entries)} to user: {entry['name']}")
        logger.info(f"  Size: {entry['size']}")
        logger.info(f"  Confidence: {entry['confidence']}")
        logger.info(f"  Link: {entry['link']}")
        logger.info(f"  Reasons:")
        for reason in entry['reasons']:
            logger.info(f"    - {reason}")
        if entry['summary']:
            logger.info(f"  Content Summary: {entry['summary'][:200]}")
        logger.info("=" * 80)

        # Print beautiful dialog box (NOT logged, only to stdout)
        # Use red color for all boxes
        RED = '\033[91m'
        RESET = '\033[0m'

        print("\n")
        print(f"{RED}╔" + "═" * 78 + "╗" + RESET)
        print(format_box_line(f"File {i + 1}/{len(entries)} - {entry['confidence']} CONFIDENCE", color_code=RED))
        print(format_box_separator("═", color_code=RED))

        # Show full path if available, otherwise just show name
        if entry.get('path'):
            # Show full path
            path = entry['path']
            if len(path) <= 68:
                print(format_box_line(f"Path: {path}", color_code=RED))
            else:
                # Wrap long paths
                print(format_box_line(f"Path: {path[:68]}", color_code=RED))
                remaining = path[68:]
                while remaining:
                    print(format_box_line(f"      {remaining[:68]}", color_code=RED))
                    remaining = remaining[68:]
        else:
            # Show just name for non-duplicates
            print(format_box_line(f"Name: {entry['name'][:68]}", color_code=RED))

        print(format_box_line(f"Size: {entry['size']}", color_code=RED))
        print(format_box_separator("─", color_code=RED))

        # Print reasons
        print(format_box_line("Reasons:", color_code=RED))
        for reason in entry['reasons']:
            # Wrap long reasons
            reason_lines = []
            reason_text = f"  • {reason}"
            if len(reason_text) <= BOX_MAX_TEXT_WIDTH:
                reason_lines.append(reason_text)
            else:
                # Word wrap
                words = reason.split()
                current_line = "  • "
                for word in words:
                    if len(current_line + word + " ") <= BOX_MAX_TEXT_WIDTH:
                        current_line += word + " "
                    else:
                        reason_lines.append(current_line.rstrip())
                        current_line = "    " + word + " "
                if current_line.strip():
                    reason_lines.append(current_line.rstrip())

            for line in reason_lines:
                print(format_box_line(line, color_code=RED))

        # Print summary if available
        if entry['summary']:
            print(format_box_separator("─", color_code=RED))
            print(format_box_line("Content Summary:", color_code=RED))

            # Clean summary text (remove BOM and other invisible characters)
            import unicodedata
            cleaned_summary = entry['summary'].replace('\ufeff', '').replace('\u200b', '')
            # Remove other control/format characters
            cleaned_summary = ''.join(
                char for char in cleaned_summary
                if unicodedata.category(char) not in ['Cf']  # Format characters
            )

            summary_lines = cleaned_summary.split('\n')
            for line in summary_lines:
                # Wrap long lines
                if len(line) <= BOX_CONTENT_MAX_WIDTH:
                    print(format_box_line(f"  {line}", color_code=RED))
                else:
                    # Word wrap
                    words = line.split()
                    current_line = "  "
                    for word in words:
                        if len(current_line + word + " ") <= BOX_MAX_TEXT_WIDTH:
                            current_line += word + " "
                        else:
                            print(format_box_line(current_line.rstrip(), color_code=RED))
                            current_line = "  " + word + " "
                    if current_line.strip():
                        print(format_box_line(current_line.rstrip(), color_code=RED))

        # Print options
        print(format_box_separator("═", color_code=RED))
        print(format_box_line("Choose action:", color_code=RED))
        print(format_box_line("  (1) Delete  |  (2) Browser  |  (3) Skip  |  (4) Next  |  (q) Quit", color_code=RED))
        print(f"{RED}╚" + "═" * 78 + "╝" + RESET)
        print("Your choice: ", end='', flush=True)

        choice = get_single_key().lower()
        print(choice)  # Echo the choice

        # Log user choice
        logger.info(f"User choice: {choice}")

        # Process the choice
        action_complete = False
        while not action_complete:
            if choice == '1':
                # Delete file
                if not entry['file_id']:
                    logger.error("Cannot delete: File ID not found")
                    logger.error("This should not happen - please report as bug")
                    break

                logger.debug(f"Attempting to move file to trash: {entry['file_id']}")
                try:
                    # Move to trash instead of permanent delete
                    service.files().update(
                        fileId=entry['file_id'],
                        body={'trashed': True}
                    ).execute()
                    logger.info(f"✅ Moved to trash: {entry['name']}")
                    print(f"✅ Moved to trash successfully!")
                    logger.debug(f"Freed up {entry['size']}")
                    log_deleted_file(folder_id, entry['name'], entry['link'], entry['size'])
                    deleted_files.add(entry['file_id'])
                    deleted_count += 1
                    action_complete = True
                except HttpError as e:
                    if e.resp.status == 404:
                        # File already deleted or doesn't exist
                        logger.warning(f"⚠️  File not found (404) - treating as deleted: {entry['name']}")
                        print(f"⚠️  File not found (404) - treating as deleted")
                        logger.debug("File may have been deleted by another process or user")
                        log_deleted_file(folder_id, entry['name'], entry['link'], entry['size'])
                        deleted_files.add(entry['file_id'])
                        deleted_count += 1
                        action_complete = True
                    else:
                        logger.error(f"❌ Failed to delete: HTTP {e.resp.status}")
                        logger.error(f"Error details: {e}")
                        print(f"❌ Failed to delete: {e}")
                        logger.warning("Skipping this file - check permissions or try again later")
                        action_complete = True
                except Exception as e:
                    logger.error(f"❌ Failed to delete: {type(e).__name__}")
                    logger.error(f"Error details: {e}")
                    print(f"❌ Failed to delete: {e}")
                    logger.warning("Unexpected error - skipping this file")
                    action_complete = True

            elif choice == '2':
                # Open in browser
                logger.info(f"🌐 Opening in browser: {entry['link']}")
                print(f"\n🌐 Opening in browser...")
                try:
                    webbrowser.open(entry['link'])
                    print("File opened in browser. Please review and choose an action.")
                    print("\nYour choice: ", end='', flush=True)
                    choice = get_single_key().lower()
                    print(choice)
                    logger.info(f"After browser review, user choice: {choice}")
                    # Continue loop with new choice
                except Exception as e:
                    logger.error(f"Failed to open browser: {e}")
                    print(f"❌ Failed to open browser: {e}")
                    action_complete = True

            elif choice == '3':
                # Skip
                logger.info(f"⏭️  Skipped: {entry['name']}")
                print(f"⏭️  Skipped")
                log_skipped_file(folder_id, entry['name'], entry['link'], entry['size'])
                skipped_files.add(entry['file_id'])
                skipped_count += 1
                action_complete = True

            elif choice == '4':
                # Next - just move to next file without logging
                logger.info(f"➡️  Next: {entry['name']}")
                print(f"➡️  Moving to next file")
                action_complete = True

            elif choice == 'q':
                print("\n" + "═" * 80)
                print("CLEANUP SESSION SUMMARY")
                print("═" * 80)
                logger.info("")
                logger.info("=" * 80)
                logger.info("CLEANUP SESSION SUMMARY")
                logger.info("=" * 80)
                logger.info(f"Files deleted: {deleted_count}")
                logger.info(f"Files skipped: {skipped_count}")
                logger.info(f"Files already processed: {already_processed_count}")
                logger.info(f"Files remaining: {len(entries) - i - 1}")
                print(f"Files deleted: {deleted_count}")
                print(f"Files skipped: {skipped_count}")
                print(f"Files already processed: {already_processed_count}")
                print(f"Files remaining: {len(entries) - i - 1}")
                logger.info("")
                logger.info(f"Logs saved:")
                logger.info(f"  Deleted: {deleted_log}")
                logger.info(f"  Skipped: {skipped_log}")
                print(f"\nLogs saved:")
                print(f"  Deleted: {deleted_log}")
                print(f"  Skipped: {skipped_log}")

                # Colored tip box advising to rerun for empty folder detection
                print_colored_tip_box([
                    "💡 TIP: Rerun the script to detect newly empty folders!",
                    "",
                    "Deleting files may have created empty folders that can now",
                    "be removed. Run the script again to detect them."
                ])

                print("═" * 80)
                return

            else:
                logger.warning(f"Invalid choice: {choice}. Please press 1, 2, 3, 4, or q")
                print(f"❌ Invalid choice '{choice}'. Please press 1, 2, 3, 4, or q")
                print("Your choice: ", end='', flush=True)
                choice = get_single_key().lower()
                print(choice)
                logger.info(f"Retrying with choice: {choice}")

        print("")  # Add spacing after action

    print("\n" + "═" * 80)
    print("CLEANUP SESSION COMPLETE")
    print("═" * 80)
    logger.info("=" * 80)
    logger.info("CLEANUP SESSION COMPLETE")
    logger.info("=" * 80)
    logger.info(f"Files deleted: {deleted_count}")
    logger.info(f"Files skipped: {skipped_count}")
    logger.info(f"Files already processed: {already_processed_count}")
    print(f"Files deleted: {deleted_count}")
    print(f"Files skipped: {skipped_count}")
    print(f"Files already processed: {already_processed_count}")
    logger.info("")
    logger.info(f"Logs saved:")
    logger.info(f"  Deleted: {deleted_log}")
    logger.info(f"  Skipped: {skipped_log}")
    print(f"\nLogs saved:")
    print(f"  Deleted: {deleted_log}")
    print(f"  Skipped: {skipped_log}")

    # Colored tip box advising to rerun for empty folder detection
    print_colored_tip_box([
        "💡 TIP: Rerun the script to detect newly empty folders!",
        "",
        "Deleting files may have created empty folders that can now",
        "be removed. Run the script again to detect them."
    ])

    print("═" * 80 + "\n")

# ============================================================================
# MAIN
# ============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Analyze Google Drive folder for cleanup candidates',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # Full workflow: Analyze AND clean a folder (default with Claude)
  python analyze_drive_cleanup.py "https://drive.google.com/drive/u/0/folders/1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh"

  # Only analyze (no interactive cleanup)
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --analyze

  # Only cleanup (interactive mode with existing report)
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --clean

  # Analyze without Claude (faster, simpler summaries)
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --analyze --no-claude

  # Use different AWS profile and region
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --aws-profile prod --aws-region eu-central-1

  # Set minimum age threshold to 30 days
  python analyze_drive_cleanup.py 1tW34LrY4e1e3OIMJkBljP0JS5atcYXJh --analyze --min-age-days 30

  # Analyze entire Drive (no folder parameter)
  python analyze_drive_cleanup.py --analyze
        '''
    )

    parser.add_argument(
        'folder',
        nargs='?',
        default=None,
        help='Google Drive folder URL or ID (required for most operations)'
    )

    parser.add_argument(
        '--analyze',
        action='store_true',
        help='Run analysis only (without interactive cleanup)'
    )

    parser.add_argument(
        '--clean',
        action='store_true',
        help='Run interactive cleanup only (without analysis)'
    )

    parser.add_argument(
        '--max-files',
        type=int,
        default=10000,
        help='Maximum number of files to scan (default: 10000)'
    )

    parser.add_argument(
        '--no-claude',
        action='store_true',
        help='Disable Claude AI summaries (use simple text extraction instead)'
    )

    parser.add_argument(
        '--aws-profile',
        type=str,
        default='dev',
        help='AWS profile to use for Bedrock (default: dev)'
    )

    parser.add_argument(
        '--aws-region',
        type=str,
        default='us-east-1',
        help='AWS region for Bedrock (default: us-east-1)'
    )

    parser.add_argument(
        '--min-age-days',
        type=int,
        default=90,
        help='Minimum age in days for files to be considered for deletion (default: 90)'
    )

    args = parser.parse_args()

    # Record start time for logging
    start_time = datetime.now(timezone.utc)

    logger.info("Google Drive Cleanup Analyzer")
    logger.info("=" * 80)

    # Extract folder ID if provided
    folder_id = None
    if args.folder:
        folder_id = extract_folder_id(args.folder)
        logger.info(f"Target folder: {folder_id}")
    else:
        logger.info("Target: Entire Google Drive")
        folder_id = "full_drive"  # Use default for logging

    # Set up file logging
    setup_file_logging(folder_id, start_time)

    # Determine which operations to run
    # If no flags specified, run both analyze and clean
    if not args.analyze and not args.clean:
        run_analyze = True
        run_clean = True
        logger.info("Mode: Full workflow (Analyze + Clean)")
    else:
        run_analyze = args.analyze
        run_clean = args.clean

        steps = []
        if run_analyze:
            steps.append("Analyze")
        if run_clean:
            steps.append("Clean")
        logger.info(f"Mode: {' + '.join(steps)}")

    logger.info("")

    # ========================================================================
    # ANALYSIS PHASE
    # ========================================================================
    if run_analyze:
        logger.info("=" * 80)
        logger.info("PHASE 1: ANALYSIS")
        logger.info("=" * 80)

        # Authenticate (read-only for analysis)
        logger.info("Authenticating with Google Drive API...")
        service = authenticate(write_access=False)
        logger.info("Authentication successful!")

        # Check if Claude should be used (default: yes, unless --no-claude)
        use_claude = not args.no_claude

        if use_claude:
            logger.info("Content analysis enabled with Claude AI")
            if not HAS_BEDROCK:
                logger.error("boto3 package not installed. Install with: pip install boto3")
                logger.error("Or use --no-claude for simple summaries")
                sys.exit(1)
            logger.info(f"AWS profile: {args.aws_profile}")
            logger.info(f"AWS region: {args.aws_region}")
        else:
            logger.info("Content analysis enabled (simple mode, no Claude)")

        # Check extraction libraries
        if not HAS_PDF:
            logger.warning("PyPDF2 not installed - PDF analysis will be skipped")
        if not HAS_DOCX:
            logger.warning("python-docx not installed - Word document analysis will be skipped")
        if not HAS_EXCEL:
            logger.warning("openpyxl not installed - Excel analysis will be skipped")

        # Load previously skipped files to avoid re-analyzing them
        _, skipped_files = load_processed_files(folder_id)
        if skipped_files:
            logger.info(f"Loaded {len(skipped_files)} previously skipped files - will skip content analysis for these")

        # Create analyzer
        analyzer = FileAnalyzer(
            service,
            analyze_content=True,  # Always analyze content
            use_claude=use_claude,
            aws_profile=args.aws_profile,
            aws_region=args.aws_region,
            min_age_days=args.min_age_days,
            skipped_files=skipped_files
        )
        logger.info(f"Minimum age threshold: {args.min_age_days} days")

        # Scan drive (folder-specific or entire drive)
        if folder_id:
            # Scan the specific folder in detail
            analyzer.scan_folder(folder_id, max_files=args.max_files)
        else:
            analyzer.scan_drive(max_files=args.max_files)

        # Analyze files
        analyzer.analyze_files()

        # Analyze folders
        analyzer.analyze_empty_folders()

        # Generate report
        logger.debug("Generating analysis report")
        report = analyzer.generate_report()

        # Save report
        folder_suffix = f"_{folder_id}" if folder_id else "_full_drive"
        report_filename = os.path.join(REPORTS_DIR, f"drive_cleanup_report{folder_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")

        logger.debug(f"Writing report to {report_filename}")
        with open(report_filename, 'w', encoding='utf-8') as f:
            f.write(report)

        logger.info("")
        logger.info(f"✅ Report saved to: {report_filename}")
        logger.info("")

    # ========================================================================
    # CLEANUP PHASE
    # ========================================================================
    if run_clean:
        logger.info("=" * 80)
        logger.info("PHASE 2: INTERACTIVE CLEANUP")
        logger.info("=" * 80)

        # Find the report file
        if not folder_id:
            logger.error("Folder ID/URL is required for cleanup mode")
            sys.exit(1)

        report_file = find_latest_report(folder_id)

        if not report_file:
            logger.error(f"No cleanup report found for folder {folder_id}")
            logger.error(f"Expected file pattern: drive_cleanup_report_{folder_id}_*.json")
            logger.error("Please run analysis first (without --clean flag)")
            sys.exit(1)

        logger.info(f"Using report: {report_file}")
        logger.info("")

        # Authenticate with write access
        logger.info("Authenticating with Google Drive API (write access)...")
        service_write = authenticate(write_access=True)
        logger.info("Authentication successful!")
        logger.info("")

        # Run interactive cleanup
        interactive_cleanup(service_write, report_file, folder_id)

if __name__ == '__main__':
    main()
